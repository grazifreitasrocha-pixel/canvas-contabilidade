import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Dados extraidos do PDF Senhor Contabil
senhor = [
    ("DECORE", "Comprovante de Rendimentos do Socio", "368,00"),
    ("Alteracao contratual", "Alteracao Contratual (exceto alteracao de natureza)", "728,00"),
    ("Alteracao de natureza juridica (LTDA para E.I. e EIRELI)", "Serao duas alteracoes concomitantemente", "848,00"),
    ("Alteracao Contratual de MEI para ME Simples", "Prazo: 60 a 90 dias para conclusao do processo", "518,00"),
    ("Envio de processos trabalhistas", "Envio de processos trabalhistas eventos e-Social", "128,00"),
    ("Alteracao Receita Federal", "Alterar nome Fantasia / Telefone ou email no CNPJ", "118,00"),
    ("Abertura de empresa sem vinculo como cliente", "Abrir empresa sem vinculo de contrato de 12 meses", "848,00"),
    ("Inscricao municipal", "Empresa ja tem o CNPJ e precisa somente a inscricao municipal", "248,00"),
    ("Regularizacao de Inscricao Estadual", "Regularizacao de inscricao estadual", "128,00"),
    ("Obtencao de Alvaras", "Pedido de alvara — servicos presenciais por conta do cliente", "248,00"),
    ("Regularizacao de Alvaras", "Regularizacao de alvara — servicos presenciais por conta do cliente", "118,00"),
    ("Baixa de empresa", "Inclui toda documentacao necessaria para orgaos, nao inclui taxas e despesas com correio", "848,00"),
    ("Baixa de empresa com menos de 6 meses", "Encerrar empresa com menos de 6 meses de contrato", "1.208,00"),
    ("Comprovante de Faturamento", "Relacao de faturamento assinada pelo contador", "68,00"),
    ("Declaracao de previsao de faturamento", "Para aberturas de contas bancarias", "68,00"),
    ("Formularios diversos (bancos, fornecedores)", "Preenchido e assinado pelo contador — custo dos Correios cobrado a parte", "58,00"),
    ("Consultoria tributaria", "Analise tributaria para buscar melhorias", "608,00"),
    ("Escrituracao fiscal via papel", "Envio das escrituracoes impressas e assinadas — Registro em Cartorio por conta do cliente", "248,00"),
    ("Reemissao de guia de imposto", "Emitir nova guia: GPS, TFE, DAS, DARF, IRRF, parcela de GPS/DAS, divida ativa", "14,00"),
    ("Reemissao de guia do FGTS", "Emitir nova guia referente ao FGTS", "32,00"),
    ("Reemissao de guia do GRRF", "Emitir nova guia referente ao GRRF", "42,00"),
    ("Envio de GFIP / DCTFWEB - Informacoes para o INSS", "Envio de GFIP/DCTFWEB para o INSS", "88,00"),
    ("Compensacao de GPS", "Compensar GPS pago em duplicidade ou valor maior", "68,00"),
    ("Compensacao de DAS", "Compensar DAS pago em duplicidade ou valor maior", "68,00"),
    ("Redarf - Retificacao para DARF", "Retificacao para DARF pago em duplicidade ou com informacao errada", "48,00"),
    ("Alteracao Folha de pagamento", "Alterar algum dado da folha de pagamento apos fechamento", "78,00"),
    ("Conversao de Extratos (PDF para OFX)", "Converter PDF em arquivo OFX (valor por competencia)", "52,00"),
    ("Emissao de CND", "Emissao de CND — CND presencial nao realizada", "68,00"),
    ("Emissao de RPA", "Emissao de Recibo de Pagamento de Autonomo", "68,00"),
    ("Entrega de declaracoes (periodos anteriores)", "RAIS, DEFIS, DIRF, DMOB, DMED, ECD, ECF — periodos anteriores", "158,00"),
    ("Verificacao de pendencias e quitacao de debitos", "Levantamento de pendencias e emissao das guias — somente online", "38,00"),
    ("Parcelamento de impostos (REFIS/RFB/PGFN/SN)", "Verificar pendencias e formalizar parcelamento — somente online", "158,00"),
    ("CPOM / CEPOM", "Preenchimento dos formularios", "248,00"),
    ("Retificacao de Declaracao (REINF)", "R$ 78,00 por trimestre retificado individualmente", "78,00"),
    ("Elaboracao de indices financeiros", "Elaboracao de indices: liquidez, FDC, EBITDA", "138,00"),
    ("Ausencia de emissao de Nota Fiscal (DANFE)", "Declaracao de ausencia de emissao de Nota Fiscal", "38,00"),
    ("Reabertura de periodo de apuracao", "Reabertura para importacao de notas", "48,00"),
    ("Cadastro de procuracao para terceiros", "Cadastro de procuracao na Receita Federal para empresas terceiras", "38,00"),
    ("Reabertura de movimentacoes", "Reabertura da aba de movimentacoes para inclusao de lancamentos", "48,00"),
    ("Envio de informacao E-Social SST com XML", "Envio das informacoes referentes ao E-Social SST com XML", "48,00"),
    ("Envio de informacao E-Social SST sem XML", "Envio das informacoes referentes ao E-Social SST sem XML", "68,00"),
    ("Certificado Digital A1 e-CPF", "Certificado Digital A1 e-CPF", "155,00"),
    ("Certificado Digital A1 e-CNPJ", "Certificado Digital A1 e-CNPJ", "188,00"),
    ("Emissao de Ata de Lucros", "Emissao de Ata de Lucros", "98,00"),
]

# Cores
COR_VERDE     = "25695C"
COR_VERDE_ESC = "144137"
COR_DOURADO   = "D2AE6D"
COR_BRANCO    = "FFFFFF"
COR_CINZA_F   = "F5F9F7"
COR_TEXTO     = "1A2E2A"
COR_CINZA_M   = "6B7280"

# Cor identidade Senhor Contabil (azul escuro para diferenciar)
COR_SC_HEADER = "1A3A5C"
COR_SC_LINHA  = "EBF2FA"
COR_SC_VALOR  = "1A3A5C"

arquivo = r"C:\Users\grazi.RAPHAEL\Downloads\Robo\Canvas_Tabela_Servicos.xlsx"
wb = openpyxl.load_workbook(arquivo)

# ---- ABA 1: renomear ----
ws1 = wb.active
ws1.title = "Canvas Contabilidade"

# ---- ABA 2: Senhor Contabil ----
ws2 = wb.create_sheet("Senhor Contabil - Referencia")

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 45
ws2.column_dimensions["C"].width = 38
ws2.column_dimensions["D"].width = 18

borda_linha = Border(
    bottom=Side(style="thin", color="D0DFF0"),
    left=Side(style="thin",   color="D0DFF0"),
    right=Side(style="thin",  color="D0DFF0"),
)
borda_header = Border(
    bottom=Side(style="medium", color=COR_DOURADO),
    top=Side(style="thin", color=COR_SC_HEADER),
)

# Linha 1: espaco
ws2.row_dimensions[1].height = 10

# Linha 2: titulo
ws2.merge_cells("A2:D2")
ws2.row_dimensions[2].height = 38
c2 = ws2["A2"]
c2.value = "SENHOR CONTABIL — REFERENCIA DE MERCADO"
c2.font = Font(name="Calibri", bold=True, size=14, color=COR_BRANCO)
c2.fill = PatternFill("solid", fgColor=COR_SC_HEADER)
c2.alignment = Alignment(horizontal="center", vertical="center")

# Linha 3: subtitulo
ws2.merge_cells("A3:D3")
ws2.row_dimensions[3].height = 20
c3 = ws2["A3"]
c3.value = "Tabela de servicos extraida do PDF Senhor Contabil — usada como referencia de mercado"
c3.font = Font(name="Calibri", italic=True, size=9, color=COR_SC_HEADER)
c3.fill = PatternFill("solid", fgColor="E8F0FB")
c3.alignment = Alignment(horizontal="center", vertical="center")

# Linha 4: espaco
ws2.row_dimensions[4].height = 6

# Linha 5: cabecalho colunas
ws2.row_dimensions[5].height = 26
hdrs = ["#", "SERVICO", "DESCRICAO", "VALOR (R$)"]
cols = ["A", "B", "C", "D"]
for col, hdr in zip(cols, hdrs):
    c = ws2[f"{col}5"]
    c.value = hdr
    c.font = Font(name="Calibri", bold=True, size=10, color=COR_BRANCO)
    c.fill = PatternFill("solid", fgColor=COR_SC_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = borda_header

# Dados
for idx, (servico, descricao, valor) in enumerate(senhor):
    row = idx + 6
    ws2.row_dimensions[row].height = 18
    fill_cor = COR_BRANCO if idx % 2 == 0 else "EBF2FA"

    ca = ws2[f"A{row}"]
    ca.value = idx + 1
    ca.font = Font(name="Calibri", size=9, color="8BAFD0", bold=True)
    ca.fill = PatternFill("solid", fgColor=fill_cor)
    ca.alignment = Alignment(horizontal="center", vertical="center")
    ca.border = borda_linha

    cb = ws2[f"B{row}"]
    cb.value = servico
    cb.font = Font(name="Calibri", size=10, color=COR_TEXTO, bold=True)
    cb.fill = PatternFill("solid", fgColor=fill_cor)
    cb.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    cb.border = borda_linha

    cc = ws2[f"C{row}"]
    cc.value = descricao
    cc.font = Font(name="Calibri", size=9, color=COR_CINZA_M)
    cc.fill = PatternFill("solid", fgColor=fill_cor)
    cc.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    cc.border = borda_linha

    cd = ws2[f"D{row}"]
    cd.value = valor
    cd.font = Font(name="Calibri", size=10, bold=True, color=COR_SC_VALOR)
    cd.fill = PatternFill("solid", fgColor=fill_cor)
    cd.alignment = Alignment(horizontal="center", vertical="center")
    cd.border = borda_linha

# Rodape
last = len(senhor) + 6
ws2.merge_cells(f"A{last}:D{last}")
ws2.row_dimensions[last].height = 18
cr = ws2[f"A{last}"]
cr.value = "Fonte: Senhor Contabil — Contabilidade Online   |   Referencia de mercado para precificacao"
cr.font = Font(name="Calibri", bold=True, size=9, color=COR_BRANCO)
cr.fill = PatternFill("solid", fgColor=COR_SC_HEADER)
cr.alignment = Alignment(horizontal="center", vertical="center")
for col in ["A","B","C","D"]:
    ws2[f"{col}{last}"].border = Border(top=Side(style="medium", color=COR_DOURADO))

ws2.freeze_panes = "B6"

# ---- ABA 3: Comparativo ----
ws3 = wb.create_sheet("Comparativo")

ws3.column_dimensions["A"].width = 6
ws3.column_dimensions["B"].width = 42
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 18
ws3.column_dimensions["E"].width = 18

# Titulo
ws3.row_dimensions[1].height = 10
ws3.merge_cells("A2:E2")
ws3.row_dimensions[2].height = 38
ct = ws3["A2"]
ct.value = "COMPARATIVO DE PRECOS — CANVAS vs SENHOR CONTABIL"
ct.font = Font(name="Calibri", bold=True, size=14, color=COR_BRANCO)
ct.fill = PatternFill("solid", fgColor=COR_VERDE_ESC)
ct.alignment = Alignment(horizontal="center", vertical="center")

ws3.merge_cells("A3:E3")
ws3.row_dimensions[3].height = 20
cs = ws3["A3"]
cs.value = "Servicos em comum entre Canvas Contabilidade e Senhor Contabil com seus respectivos valores"
cs.font = Font(name="Calibri", italic=True, size=9, color=COR_VERDE)
cs.fill = PatternFill("solid", fgColor="EBF5F0")
cs.alignment = Alignment(horizontal="center", vertical="center")

ws3.row_dimensions[4].height = 6

# Cabecalho
ws3.row_dimensions[5].height = 26
hdrs3 = ["#", "SERVICO", "CANVAS", "SENHOR CONTABIL", "DIFERENCA"]
cols3 = ["A","B","C","D","E"]
for col, hdr in zip(cols3, hdrs3):
    c = ws3[f"{col}5"]
    c.value = hdr
    c.font = Font(name="Calibri", bold=True, size=10, color=COR_BRANCO)
    c.fill = PatternFill("solid", fgColor=COR_VERDE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(bottom=Side(style="medium", color=COR_DOURADO))

# Comparativos mapeados manualmente
comparativos = [
    ("Abertura de empresa", "1.100,00", "848,00"),
    ("Alteracao contratual", "880,00", "728,00"),
    ("Baixa de empresa", "550,00", "848,00"),
    ("Alvaras / Obtencao de Alvaras", "297,00 - 300,00", "248,00"),
    ("Certificado Digital", "180,00", "155,00 (CPF) / 188,00 (CNPJ)"),
    ("Consultoria / Analise tributaria", "550,00", "608,00"),
    ("Parcelamento tributario", "38,21", "158,00"),
    ("Compensacao de creditos (GPS/DAS)", "A consultar", "68,00"),
    ("Retificacao / Regularizacao", "400,00", "78,00 - 128,00"),
    ("Emissao de RPA", "A consultar", "68,00"),
    ("Emissao de CND", "A consultar", "68,00"),
    ("Emissao de Ata de Lucros", "A consultar", "98,00"),
    ("Envio de processos trabalhistas", "A consultar", "128,00"),
    ("Folha Domestica / Alteracao Folha", "180,00", "78,00"),
    ("DECORE", "A consultar", "368,00"),
]

for idx, (srv, canvas_v, sc_v) in enumerate(comparativos):
    row = idx + 6
    ws3.row_dimensions[row].height = 20
    fill_cor = COR_BRANCO if idx % 2 == 0 else COR_CINZA_F

    bln = Border(bottom=Side(style="thin", color="D0E5DC"), left=Side(style="thin", color="D0E5DC"), right=Side(style="thin", color="D0E5DC"))

    ca = ws3[f"A{row}"]
    ca.value = idx + 1
    ca.font = Font(name="Calibri", size=9, color="8FB8AD", bold=True)
    ca.fill = PatternFill("solid", fgColor=fill_cor)
    ca.alignment = Alignment(horizontal="center", vertical="center")
    ca.border = bln

    cb = ws3[f"B{row}"]
    cb.value = srv
    cb.font = Font(name="Calibri", size=10, color=COR_TEXTO)
    cb.fill = PatternFill("solid", fgColor=fill_cor)
    cb.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    cb.border = bln

    cc = ws3[f"C{row}"]
    cc.value = canvas_v
    cc.font = Font(name="Calibri", size=10, bold=True, color=COR_VERDE)
    cc.fill = PatternFill("solid", fgColor=fill_cor)
    cc.alignment = Alignment(horizontal="center", vertical="center")
    cc.border = bln

    cd = ws3[f"D{row}"]
    cd.value = sc_v
    cd.font = Font(name="Calibri", size=10, bold=True, color=COR_SC_VALOR)
    cd.fill = PatternFill("solid", fgColor=fill_cor)
    cd.alignment = Alignment(horizontal="center", vertical="center")
    cd.border = bln

    ce = ws3[f"E{row}"]
    ce.value = "Ver detalhes"
    ce.font = Font(name="Calibri", size=9, italic=True, color=COR_CINZA_M)
    ce.fill = PatternFill("solid", fgColor=fill_cor)
    ce.alignment = Alignment(horizontal="center", vertical="center")
    ce.border = bln

# Rodape comparativo
last3 = len(comparativos) + 6
ws3.merge_cells(f"A{last3}:E{last3}")
ws3.row_dimensions[last3].height = 18
cr3 = ws3[f"A{last3}"]
cr3.value = "Canvas Contabilidade  |  Referencia de mercado — Senhor Contabil"
cr3.font = Font(name="Calibri", bold=True, size=9, color=COR_BRANCO)
cr3.fill = PatternFill("solid", fgColor=COR_VERDE_ESC)
cr3.alignment = Alignment(horizontal="center", vertical="center")
for col in ["A","B","C","D","E"]:
    ws3[f"{col}{last3}"].border = Border(top=Side(style="medium", color=COR_DOURADO))

ws3.freeze_panes = "B6"

wb.save(arquivo)
print(f"Salvo: {arquivo}")
print(f"Abas: {[s.title for s in wb.worksheets]}")
print(f"Servicos Senhor Contabil adicionados: {len(senhor)}")
print(f"Comparativos gerados: {len(comparativos)}")
