#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Planilha Comparativo Reforma Tributária: PIS/COFINS → CBS
LCP 214/2025 — Gerada automaticamente
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────── PALETA DE CORES ────────────────────────────
COR = {
    "AZUL_ESC":  "1F3864", "AZUL_MED":  "2E75B6", "AZUL_CLA":  "BDD7EE",
    "VRD_ESC":   "375623", "VRD_CLA":   "E2EFDA",
    "MARROM":    "833C00", "AMARELO":   "FFF2CC", "LARANJA":   "FCE4D6",
    "ROXO":      "7030A0", "CINZA":     "F2F2F2", "CINZA_ESC": "D9D9D9",
    "INPUT":     "DAEEF3", "RESULT":    "C6EFCE", "VERM":      "FFC7CE",
    "BRANCO":    "FFFFFF", "VERDE_N":   "00B050", "VERM_N":    "FF0000",
}
BRL = 'R$ #,##0.00'
PCT = '0.00%'

# ─────────────────────────── HELPERS ───────────────────────────────────

def bd(estilo="thin"):
    s = Side(style=estilo)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, valor=None, fmt=None, bg=None, bold=False, italic=False,
         cor_txt="000000", sz=11, halign="left", wrap=True, span=None):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(name="Calibri", bold=bold, italic=italic, color=cor_txt, size=sz)
    if bg:
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    c.border = bd()
    if fmt:
        c.number_format = fmt
    if span:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span)
    return c

def rh(ws, row, h):   ws.row_dimensions[row].height = h
def cw(ws, col, w):   ws.column_dimensions[col].width = w

def titulo(ws, row, texto, span=6, bg="AZUL_ESC", sz=13):
    cell(ws, row, 1, texto, bg=COR[bg], bold=True, cor_txt="FFFFFF",
         sz=sz, halign="center", span=span)

def secao(ws, row, texto, span=6, bg="AZUL_MED"):
    cell(ws, row, 1, texto, bg=COR[bg], bold=True, cor_txt="FFFFFF",
         sz=11, halign="center", span=span)

def nota(ws, row, texto, bg="FFF9F0"):
    c = ws.cell(row=row, column=1, value=f"• {texto}")
    c.font = Font(name="Calibri", italic=True, size=10, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.border = bd()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    rh(ws, row, 22)


# ═══════════════════════════════════════════════════════════════════════
#   SHEET 1 — LUCRO PRESUMIDO
# ═══════════════════════════════════════════════════════════════════════
def sheet_lucro_presumido(wb):
    ws = wb.create_sheet("Lucro Presumido")
    for col, w in zip("ABCDEF", [44, 20, 13, 20, 13, 20]):
        cw(ws, col, w)

    # ── Título
    rh(ws, 1, 48)
    titulo(ws, 1, "REFORMA TRIBUTÁRIA — LUCRO PRESUMIDO (Regime Cumulativo)\n"
                  "Comparativo: PIS + COFINS  ×  CBS  |  LCP 214/2025", sz=13)
    rh(ws, 2, 20)
    cell(ws, 2, 1, "🔵 Células AZUIS = preencha com seus dados  |  Células CINZAS = calculadas automaticamente",
         bg=COR["AZUL_CLA"], italic=True, sz=10, halign="center", span=6)
    rh(ws, 3, 8)

    # ── Inputs
    rh(ws, 4, 28)
    secao(ws, 4, "DADOS DE ENTRADA — altere os valores azuis e a planilha recalcula automaticamente")

    entradas = [
        (5,  "Custo de aquisição / produção (R$)",                  1000.00, BRL,
              "Custo total pago pelo produto ou serviço prestado"),
        (6,  "Margem de lucro desejada (% sobre o preço base)",     0.30,    PCT,
              "Ex: 30% → o lucro representa 30% do preço de venda"),
        (7,  "Alíquota CBS estimada — por fora (editável)",         0.088,   PCT,
              "Estimativa: ~8,8%. Será fixada por Resolução do Senado Federal"),
        (8,  "% do custo que teve CBS nos insumos (para crédito)",  0.70,    PCT,
              "Ex: 70% do seu custo de compra vem de fornecedores que cobraram CBS"),
    ]
    for row, lbl, val, fmt, tip in entradas:
        rh(ws, row, 22)
        cell(ws, row, 1, lbl, bg=COR["CINZA"], bold=True)
        cell(ws, row, 2, val, fmt=fmt, bg=COR["INPUT"], bold=True, cor_txt="0070C0", halign="right")
        cell(ws, row, 3, tip, bg=COR["BRANCO"], italic=True, sz=10, cor_txt="595959", span=6)
    rh(ws, 9, 8)

    # ── Cabeçalhos colunas
    rh(ws, 10, 38)
    cell(ws, 10, 1, "ITEM DE ANÁLISE", bg=COR["AZUL_ESC"], bold=True, cor_txt="FFFFFF", halign="center")
    cell(ws, 10, 2, "ANTES (até 2026)\nPIS + COFINS\n(por dentro)", bg=COR["VRD_ESC"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 10, 3, "%\ndo\nPreço", bg=COR["VRD_ESC"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 10, 4, "DEPOIS (2027+)\nCBS\n(por fora)", bg=COR["MARROM"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 10, 5, "%\ndo\nPreço", bg=COR["MARROM"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 10, 6, "VARIAÇÃO\n(Depois − Antes)", bg=COR["ROXO"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)

    rh(ws, 11, 22)
    secao(ws, 11, "FORMAÇÃO DO PREÇO E TRIBUTOS")

    # ── Linha 12: Preço base
    # ANTES: P = Custo / (1 - 0.0365 - Margem)
    # DEPOIS: PB = Custo*(1 - %cred*aliq) / (1 - aliq - Margem)
    rh(ws, 12, 28)
    cell(ws, 12, 1, "Preço de venda / Preço base (R$)", bg=COR["CINZA"], bold=True)
    cell(ws, 12, 2, "=B5/(1-0.0365-B6)",          fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, 12, 3, "=100%",                        fmt=PCT, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 12, 4, "=B5*(1-B8*B7)/(1-B7-B6)",    fmt=BRL, bg=COR["AMARELO"], bold=True, halign="right")
    cell(ws, 12, 5, "=100%",                        fmt=PCT, bg=COR["AMARELO"], halign="right")
    cell(ws, 12, 6, "=D12-B12",                     fmt=BRL, bg=COR["CINZA"], halign="right")

    # Linha 13: PIS
    rh(ws, 13, 22)
    cell(ws, 13, 1, "(-) PIS — 0,65% por dentro do preço", bg=COR["CINZA"])
    cell(ws, 13, 2, "=-B12*0.0065",  fmt=BRL, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 13, 3, "=ABS(B13)/B12", fmt=PCT, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 13, 4, "=0",            fmt=BRL, bg=COR["AMARELO"], halign="right", cor_txt="AAAAAA")
    cell(ws, 13, 5, '="n/a"',                 bg=COR["AMARELO"], halign="center", cor_txt="AAAAAA")
    cell(ws, 13, 6, "=D13-B13",      fmt=BRL, bg=COR["CINZA"],  halign="right")

    # Linha 14: COFINS
    rh(ws, 14, 22)
    cell(ws, 14, 1, "(-) COFINS — 3,00% por dentro do preço", bg=COR["CINZA"])
    cell(ws, 14, 2, "=-B12*0.03",    fmt=BRL, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 14, 3, "=ABS(B14)/B12", fmt=PCT, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 14, 4, "=0",            fmt=BRL, bg=COR["AMARELO"], halign="right", cor_txt="AAAAAA")
    cell(ws, 14, 5, '="n/a"',                 bg=COR["AMARELO"], halign="center", cor_txt="AAAAAA")
    cell(ws, 14, 6, "=D14-B14",      fmt=BRL, bg=COR["CINZA"],  halign="right")

    # Linha 15: CBS bruta (apenas DEPOIS)
    rh(ws, 15, 22)
    cell(ws, 15, 1, "(−) CBS bruta sobre preço base (por fora)", bg=COR["CINZA"])
    cell(ws, 15, 2, '="n/a"',          bg=COR["VRD_CLA"], halign="center", cor_txt="AAAAAA")
    cell(ws, 15, 3, '="n/a"',          bg=COR["VRD_CLA"], halign="center", cor_txt="AAAAAA")
    cell(ws, 15, 4, "=-D12*B7",        fmt=BRL, bg=COR["AMARELO"], halign="right")
    cell(ws, 15, 5, "=ABS(D15)/D12",   fmt=PCT, bg=COR["AMARELO"], halign="right")
    cell(ws, 15, 6, "=D15-B15",        fmt=BRL, bg=COR["CINZA"],  halign="right")

    # Linha 16: Crédito CBS (apenas DEPOIS — novidade para o cumulativo!)
    rh(ws, 16, 22)
    cell(ws, 16, 1, "(+) Crédito CBS sobre insumos  ← NOVIDADE para regime cumulativo!", bg=COR["CINZA"])
    cell(ws, 16, 2, "=0",           fmt=BRL, bg=COR["VRD_CLA"], halign="right", cor_txt="AAAAAA")
    cell(ws, 16, 3, '="—"',                  bg=COR["VRD_CLA"], halign="center", cor_txt="AAAAAA")
    cell(ws, 16, 4, "=B5*B8*B7",    fmt=BRL, bg=COR["AMARELO"], halign="right")
    cell(ws, 16, 5, "=D16/D12",     fmt=PCT, bg=COR["AMARELO"], halign="right")
    cell(ws, 16, 6, "=D16-B16",     fmt=BRL, bg=COR["CINZA"],  halign="right")

    # Linha 17: TOTAL TRIBUTO LÍQUIDO
    rh(ws, 17, 28)
    cell(ws, 17, 1, "TOTAL TRIBUTO LÍQUIDO A RECOLHER", bg=COR["LARANJA"], bold=True)
    cell(ws, 17, 2, "=ABS(B13)+ABS(B14)",    fmt=BRL, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, 17, 3, "=B17/B12",              fmt=PCT, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, 17, 4, "=ABS(D15)-D16",         fmt=BRL, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, 17, 5, "=D17/D12",              fmt=PCT, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, 17, 6, "=D17-B17",              fmt=BRL, bg=COR["CINZA"],  bold=True, halign="right")

    # Linha 18: Custo
    rh(ws, 18, 22)
    cell(ws, 18, 1, "(−) Custo de aquisição / produção", bg=COR["CINZA"])
    cell(ws, 18, 2, "=-B5",  fmt=BRL, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 18, 3, "=B5/B12", fmt=PCT, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 18, 4, "=-B5",  fmt=BRL, bg=COR["AMARELO"], halign="right")
    cell(ws, 18, 5, "=B5/D12", fmt=PCT, bg=COR["AMARELO"], halign="right")
    cell(ws, 18, 6, "=0",    fmt=BRL, bg=COR["CINZA"],  halign="right")

    # Linha 19: LUCRO
    rh(ws, 19, 30)
    cell(ws, 19, 1, "= LUCRO BRUTO DO EMPRESÁRIO (R$)", bg=COR["AZUL_CLA"], bold=True)
    cell(ws, 19, 2, "=B12-B17-B5",  fmt=BRL, bg=COR["AZUL_CLA"], bold=True, halign="right")
    cell(ws, 19, 3, "=B19/B12",     fmt=PCT, bg=COR["AZUL_CLA"], bold=True, halign="right")
    cell(ws, 19, 4, "=D12-D17-B5",  fmt=BRL, bg=COR["AZUL_CLA"], bold=True, halign="right")
    cell(ws, 19, 5, "=D19/D12",     fmt=PCT, bg=COR["AZUL_CLA"], bold=True, halign="right")
    cell(ws, 19, 6, "=D19-B19",     fmt=BRL, bg=COR["CINZA"],  bold=True, halign="right")

    rh(ws, 20, 8)

    # ── PREÇO AO CONSUMIDOR
    rh(ws, 21, 30)
    secao(ws, 21, "PREÇO FINAL AO CONSUMIDOR", bg="AZUL_ESC")

    rh(ws, 22, 22)
    cell(ws, 22, 1, "Preço base cobrado pelo vendedor", bg=COR["CINZA"], bold=True)
    cell(ws, 22, 2, "=B12",  fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, 22, 3, "",      bg=COR["VRD_CLA"])
    cell(ws, 22, 4, "=D12",  fmt=BRL, bg=COR["AMARELO"], bold=True, halign="right")
    cell(ws, 22, 5, "",      bg=COR["AMARELO"])
    cell(ws, 22, 6, "=D22-B22", fmt=BRL, bg=COR["CINZA"], halign="right")

    rh(ws, 23, 22)
    cell(ws, 23, 1, "(+) CBS destacada por fora (paga pelo consumidor, repassada ao Fisco via Split Payment)", bg=COR["CINZA"])
    cell(ws, 23, 2, "=0",         fmt=BRL, bg=COR["VRD_CLA"], halign="right", cor_txt="AAAAAA")
    cell(ws, 23, 3, "",           bg=COR["VRD_CLA"])
    cell(ws, 23, 4, "=D12*B7",    fmt=BRL, bg=COR["AMARELO"], halign="right")
    cell(ws, 23, 5, "",           bg=COR["AMARELO"])
    cell(ws, 23, 6, "=D23-B23",   fmt=BRL, bg=COR["CINZA"], halign="right")

    rh(ws, 24, 30)
    cell(ws, 24, 1, "TOTAL PAGO PELO CONSUMIDOR (R$)", bg=COR["RESULT"], bold=True)
    cell(ws, 24, 2, "=B22+B23",  fmt=BRL, bg=COR["RESULT"], bold=True, halign="right")
    cell(ws, 24, 3, "",          bg=COR["RESULT"])
    cell(ws, 24, 4, "=D22+D23",  fmt=BRL, bg=COR["RESULT"], bold=True, halign="right")
    cell(ws, 24, 5, "",          bg=COR["RESULT"])
    cell(ws, 24, 6, "=D24-B24",  fmt=BRL, bg=COR["RESULT"], bold=True, halign="right")

    rh(ws, 25, 25)
    cell(ws, 25, 1, "Variação % no preço ao consumidor", bg=COR["CINZA"], bold=True)
    cell(ws, 25, 2, "", bg=COR["CINZA"]); cell(ws, 25, 3, "", bg=COR["CINZA"])
    cell(ws, 25, 4, "", bg=COR["CINZA"]); cell(ws, 25, 5, "", bg=COR["CINZA"])
    cell(ws, 25, 6, "=(D24-B24)/B24", fmt=PCT, bg=COR["AMARELO"], bold=True, halign="right")

    rh(ws, 26, 8)

    # ── Notas
    rh(ws, 27, 22)
    secao(ws, 27, "COMO INTERPRETAR ESTES NÚMEROS", bg="MARROM")
    notas = [
        "ANTES (PIS+COFINS): Alíquota 3,65% embutida POR DENTRO do preço. Sem crédito sobre insumos (regime cumulativo). Carga menor mas sem recuperação.",
        "DEPOIS (CBS): Alíquota ~8,8% calculada POR FORA sobre o preço base. Crédito sobre insumos (NOVIDADE para o cumulativo). Bruto maior, líquido pode ser menor.",
        "SPLIT PAYMENT (Arts. 31-34, LCP 214): Banco/maquininha recolhe a CBS automaticamente no pagamento. O empresário recebe o preço base — sem fluxo manual.",
        "EMPRESA COM MUITOS INSUMOS: Crédito CBS reduz bastante a carga líquida. Pode ser FAVORECIDA em relação ao regime anterior.",
        "EMPRESA DE SERVIÇO PURO: Poucos insumos = poucos créditos. CBS líquida tende a ser MAIOR que PIS+COFINS atual. Reajuste de preço pode ser necessário.",
    ]
    for i, n in enumerate(notas):
        nota(ws, 28 + i, n)


# ═══════════════════════════════════════════════════════════════════════
#   SHEET 2 — LUCRO REAL
# ═══════════════════════════════════════════════════════════════════════
def sheet_lucro_real(wb):
    ws = wb.create_sheet("Lucro Real")
    for col, w in zip("ABCDEF", [44, 20, 13, 20, 13, 20]):
        cw(ws, col, w)

    rh(ws, 1, 48)
    titulo(ws, 1, "REFORMA TRIBUTÁRIA — LUCRO REAL (Regime Não-Cumulativo)\n"
                  "Comparativo: PIS + COFINS  ×  CBS  |  LCP 214/2025", sz=13)
    rh(ws, 2, 20)
    cell(ws, 2, 1, "🔵 Células AZUIS = preencha com seus dados  |  Células CINZAS = calculadas automaticamente",
         bg=COR["AZUL_CLA"], italic=True, sz=10, halign="center", span=6)
    rh(ws, 3, 8)

    rh(ws, 4, 28)
    secao(ws, 4, "DADOS DE ENTRADA — altere os valores azuis e a planilha recalcula automaticamente")

    entradas = [
        (5,  "Custo de aquisição / produção (R$)",                   1000.00, BRL,
              "Custo total pago pelo produto ou serviço prestado"),
        (6,  "Margem de lucro desejada (% sobre o preço base)",      0.30,    PCT,
              "Ex: 30% → o lucro representa 30% do preço de venda"),
        (7,  "Alíquota CBS estimada — por fora (editável)",          0.088,   PCT,
              "Estimativa: ~8,8%. Será fixada por Resolução do Senado Federal"),
        (8,  "% do custo que teve CBS nos insumos (para crédito CBS)",0.80,   PCT,
              "Lucro Real tende a ter percentual maior de insumos com CBS"),
        (9,  "% do custo que gerou crédito PIS+COFINS (ANTES)",      0.70,    PCT,
              "No não-cumulativo, quanto do custo gerou crédito de PIS/COFINS?"),
    ]
    for row, lbl, val, fmt, tip in entradas:
        rh(ws, row, 22)
        cell(ws, row, 1, lbl, bg=COR["CINZA"], bold=True)
        cell(ws, row, 2, val, fmt=fmt, bg=COR["INPUT"], bold=True, cor_txt="0070C0", halign="right")
        cell(ws, row, 3, tip, bg=COR["BRANCO"], italic=True, sz=10, cor_txt="595959", span=6)
    rh(ws, 10, 8)

    # Cabeçalhos
    rh(ws, 11, 38)
    cell(ws, 11, 1, "ITEM DE ANÁLISE", bg=COR["AZUL_ESC"], bold=True, cor_txt="FFFFFF", halign="center")
    cell(ws, 11, 2, "ANTES (até 2026)\nPIS (1,65%) + COFINS (7,6%)\nNão-cumulativo — com crédito",
         bg=COR["VRD_ESC"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 11, 3, "%\ndo\nPreço", bg=COR["VRD_ESC"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 11, 4, "DEPOIS (2027+)\nCBS (~8,8%)\nPor fora — com crédito",
         bg=COR["MARROM"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 11, 5, "%\ndo\nPreço", bg=COR["MARROM"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 11, 6, "VARIAÇÃO\n(Depois − Antes)", bg=COR["ROXO"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)

    rh(ws, 12, 22)
    secao(ws, 12, "FORMAÇÃO DO PREÇO E TRIBUTOS")

    # ANTES: alíquota PIS+COFINS = 9,25%, com crédito sobre %_cred × 9,25% do custo
    # P*(1 - 0.0925 - M) = Custo*(1 - 0.0925*%cred)  →  P = Custo*(1-0.0925*%cred)/(1-0.0925-M)
    # DEPOIS: PB*(1-B7-M) = Custo*(1-B7*B8)  →  PB = Custo*(1-B7*B8)/(1-B7-M)

    rh(ws, 13, 28)
    cell(ws, 13, 1, "Preço de venda / Preço base (R$)", bg=COR["CINZA"], bold=True)
    cell(ws, 13, 2, "=B5*(1-0.0925*B9)/(1-0.0925-B6)",  fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, 13, 3, "=100%",                              fmt=PCT, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 13, 4, "=B5*(1-B7*B8)/(1-B7-B6)",           fmt=BRL, bg=COR["AMARELO"], bold=True, halign="right")
    cell(ws, 13, 5, "=100%",                              fmt=PCT, bg=COR["AMARELO"], halign="right")
    cell(ws, 13, 6, "=D13-B13",                           fmt=BRL, bg=COR["CINZA"],  halign="right")

    # PIS bruto (ANTES)
    rh(ws, 14, 22)
    cell(ws, 14, 1, "PIS bruto — 1,65% sobre o preço / CBS bruta sobre preço base", bg=COR["CINZA"])
    cell(ws, 14, 2, "=B13*0.0165",    fmt=BRL, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 14, 3, "=B14/B13",       fmt=PCT, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 14, 4, "=D13*B7",        fmt=BRL, bg=COR["AMARELO"], halign="right")
    cell(ws, 14, 5, "=D14/D13",       fmt=PCT, bg=COR["AMARELO"], halign="right")
    cell(ws, 14, 6, "=D14-B14",       fmt=BRL, bg=COR["CINZA"],  halign="right")

    # COFINS bruta (ANTES)
    rh(ws, 15, 22)
    cell(ws, 15, 1, "COFINS bruta — 7,60% sobre o preço / (incluída acima na CBS bruta)", bg=COR["CINZA"])
    cell(ws, 15, 2, "=B13*0.076",     fmt=BRL, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 15, 3, "=B15/B13",       fmt=PCT, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 15, 4, '="—"',           bg=COR["AMARELO"], halign="center", cor_txt="AAAAAA")
    cell(ws, 15, 5, '="—"',           bg=COR["AMARELO"], halign="center", cor_txt="AAAAAA")
    cell(ws, 15, 6, "",               bg=COR["CINZA"])

    # Crédito PIS+COFINS (ANTES) / CBS (DEPOIS)
    rh(ws, 16, 22)
    cell(ws, 16, 1, "(−) Crédito sobre insumos: PIS+COFINS (ANTES) / CBS (DEPOIS)", bg=COR["CINZA"])
    cell(ws, 16, 2, "=-(B14+B15)*B9", fmt=BRL, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 16, 3, "=ABS(B16)/B13",  fmt=PCT, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 16, 4, "=-(B5*B8*B7)",   fmt=BRL, bg=COR["AMARELO"], halign="right")
    cell(ws, 16, 5, "=ABS(D16)/D13",  fmt=PCT, bg=COR["AMARELO"], halign="right")
    cell(ws, 16, 6, "=D16-B16",       fmt=BRL, bg=COR["CINZA"],  halign="right")

    # TOTAL TRIBUTO LÍQUIDO
    rh(ws, 17, 28)
    cell(ws, 17, 1, "TOTAL TRIBUTO LÍQUIDO A RECOLHER", bg=COR["LARANJA"], bold=True)
    cell(ws, 17, 2, "=B14+B15+B16",     fmt=BRL, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, 17, 3, "=B17/B13",         fmt=PCT, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, 17, 4, "=D14+D16",         fmt=BRL, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, 17, 5, "=D17/D13",         fmt=PCT, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, 17, 6, "=D17-B17",         fmt=BRL, bg=COR["CINZA"],  bold=True, halign="right")

    # Custo
    rh(ws, 18, 22)
    cell(ws, 18, 1, "(−) Custo de aquisição / produção", bg=COR["CINZA"])
    cell(ws, 18, 2, "=-B5", fmt=BRL, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 18, 3, "=B5/B13", fmt=PCT, bg=COR["VRD_CLA"], halign="right")
    cell(ws, 18, 4, "=-B5", fmt=BRL, bg=COR["AMARELO"], halign="right")
    cell(ws, 18, 5, "=B5/D13", fmt=PCT, bg=COR["AMARELO"], halign="right")
    cell(ws, 18, 6, "=0", fmt=BRL, bg=COR["CINZA"], halign="right")

    # LUCRO
    rh(ws, 19, 30)
    cell(ws, 19, 1, "= LUCRO BRUTO DO EMPRESÁRIO (R$)", bg=COR["AZUL_CLA"], bold=True)
    cell(ws, 19, 2, "=B13+B17+B18", fmt=BRL, bg=COR["AZUL_CLA"], bold=True, halign="right")
    cell(ws, 19, 3, "=B19/B13",     fmt=PCT, bg=COR["AZUL_CLA"], bold=True, halign="right")
    cell(ws, 19, 4, "=D13+D17+D18", fmt=BRL, bg=COR["AZUL_CLA"], bold=True, halign="right")
    cell(ws, 19, 5, "=D19/D13",     fmt=PCT, bg=COR["AZUL_CLA"], bold=True, halign="right")
    cell(ws, 19, 6, "=D19-B19",     fmt=BRL, bg=COR["CINZA"],  bold=True, halign="right")

    rh(ws, 20, 8)
    rh(ws, 21, 30)
    secao(ws, 21, "PREÇO FINAL AO CONSUMIDOR", bg="AZUL_ESC")

    rh(ws, 22, 22)
    cell(ws, 22, 1, "Preço base cobrado pelo vendedor", bg=COR["CINZA"], bold=True)
    cell(ws, 22, 2, "=B13", fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, 22, 3, "", bg=COR["VRD_CLA"])
    cell(ws, 22, 4, "=D13", fmt=BRL, bg=COR["AMARELO"], bold=True, halign="right")
    cell(ws, 22, 5, "", bg=COR["AMARELO"])
    cell(ws, 22, 6, "=D22-B22", fmt=BRL, bg=COR["CINZA"], halign="right")

    rh(ws, 23, 22)
    cell(ws, 23, 1, "(+) CBS destacada por fora (Split Payment — recolhe automático)", bg=COR["CINZA"])
    cell(ws, 23, 2, "=0", fmt=BRL, bg=COR["VRD_CLA"], halign="right", cor_txt="AAAAAA")
    cell(ws, 23, 3, "", bg=COR["VRD_CLA"])
    cell(ws, 23, 4, "=D13*B7", fmt=BRL, bg=COR["AMARELO"], halign="right")
    cell(ws, 23, 5, "", bg=COR["AMARELO"])
    cell(ws, 23, 6, "=D23-B23", fmt=BRL, bg=COR["CINZA"], halign="right")

    rh(ws, 24, 30)
    cell(ws, 24, 1, "TOTAL PAGO PELO CONSUMIDOR (R$)", bg=COR["RESULT"], bold=True)
    cell(ws, 24, 2, "=B22+B23", fmt=BRL, bg=COR["RESULT"], bold=True, halign="right")
    cell(ws, 24, 3, "", bg=COR["RESULT"])
    cell(ws, 24, 4, "=D22+D23", fmt=BRL, bg=COR["RESULT"], bold=True, halign="right")
    cell(ws, 24, 5, "", bg=COR["RESULT"])
    cell(ws, 24, 6, "=D24-B24", fmt=BRL, bg=COR["RESULT"], bold=True, halign="right")

    rh(ws, 25, 25)
    cell(ws, 25, 1, "Variação % no preço ao consumidor", bg=COR["CINZA"], bold=True)
    for c in [2,3,4,5]: cell(ws, 25, c, "", bg=COR["CINZA"])
    cell(ws, 25, 6, "=(D24-B24)/B24", fmt=PCT, bg=COR["AMARELO"], bold=True, halign="right")

    rh(ws, 26, 8)
    rh(ws, 27, 22)
    secao(ws, 27, "COMO INTERPRETAR ESTES NÚMEROS", bg="MARROM")
    notas = [
        "ANTES (PIS+COFINS não-cumulativo): Alíquota 9,25% sobre receita, COM crédito sobre insumos. Já existia o mecanismo de crédito.",
        "DEPOIS (CBS): Alíquota similar (~8,8%) mas calculada POR FORA. O crédito permanece e tende a ser mais abrangente pela LCP 214.",
        "LUCRO REAL é o regime com MENOR impacto na transição — a lógica de crédito permanece, apenas muda a forma de cálculo e recolhimento.",
        "ATENÇÃO: Durante 2026-2032, PIS/COFINS NÃO integram a base do CBS (Art.12 §2° V, LCP 214). Sem dupla tributação no período de transição.",
        "SPLIT PAYMENT garante que a CBS seja recolhida automaticamente pelo sistema financeiro, reduzindo inadimplência e simplificando o processo.",
    ]
    for i, n in enumerate(notas):
        nota(ws, 28 + i, n)


# ═══════════════════════════════════════════════════════════════════════
#   SHEET 3 — SIMPLES NACIONAL
# ═══════════════════════════════════════════════════════════════════════
def sheet_simples_nacional(wb):
    ws = wb.create_sheet("Simples Nacional")

    # Larguras das 11 colunas A-K
    for col, w in zip("ABCDEFGHIJK", [4, 24, 11, 14, 7, 7, 9, 7, 9, 13, 13]):
        cw(ws, col, w)

    # ── DADOS OFICIAIS DOS 5 ANEXOS ──────────────────────────────────────
    # (nome, header_col_j, [(aliq_nominal, deducao, irpj%, csll%, cofins%, pis%, cpp%, colj%), ...])
    # Fonte: LC 123/2006 atualizada pela LC 155/2016 e LC 167/2019
    LIMITES = [180_000, 360_000, 720_000, 1_800_000, 3_600_000, 4_800_000]

    ANEXOS = [
        ("ANEXO I — COMERCIO", "ICMS",
         [(0.0400,      0, 5.50,  3.50, 12.74, 2.76, 41.50, 34.00),
          (0.0730,  5_940, 5.50,  3.50, 12.74, 2.76, 41.50, 34.00),
          (0.0950, 13_860, 5.50,  3.50, 12.74, 2.76, 41.50, 34.00),
          (0.1070, 22_500, 5.50,  3.50, 12.74, 2.76, 41.50, 34.00),
          (0.1430, 87_300, 5.50,  3.50, 12.74, 2.76, 41.50, 34.00),
          (0.1900,378_000,13.50, 10.00, 28.27, 6.13, 42.10,  0.00)]),

        ("ANEXO II — INDUSTRIA", "ICMS+IPI",
         [(0.0450,      0, 5.50,  3.50, 11.51, 2.49, 37.50, 39.50),
          (0.0780,  5_940, 5.50,  3.50, 11.51, 2.49, 37.50, 39.50),
          (0.1000, 13_860, 5.50,  3.50, 11.51, 2.49, 37.50, 39.50),
          (0.1120, 22_500, 5.50,  3.50, 11.51, 2.49, 37.50, 39.50),
          (0.1470, 85_500, 5.50,  3.50, 11.51, 2.49, 37.50, 39.50),
          (0.3000,720_000, 8.50,  7.50, 20.96, 4.54, 23.50, 35.00)]),

        ("ANEXO III — SERVICOS (agencia, representacao, corretagem, etc.)", "ISS",
         [(0.0600,      0,  4.00,  3.50, 12.82, 2.78, 43.40, 33.50),
          (0.1120,  9_360,  4.00,  3.50, 14.05, 3.05, 43.40, 32.00),
          (0.1350, 17_640,  4.00,  3.50, 13.64, 2.96, 43.40, 32.50),
          (0.1600, 35_640,  4.00,  3.50, 13.64, 2.96, 43.40, 32.50),
          (0.2100,125_640,  4.00,  3.50, 12.82, 2.78, 43.40, 33.50),
          (0.3300,648_000, 35.00, 15.00, 16.03, 3.47, 30.50,  0.00)]),

        ("ANEXO IV — SERVICOS (construcao, vigilancia, limpeza — sem CPP no DAS)", "ISS",
         [(0.0450,      0, 18.80, 15.20, 17.67, 3.83,  0.00, 44.50),
          (0.0900,  8_100, 19.80, 15.20, 20.55, 4.45,  0.00, 40.00),
          (0.1020, 12_420, 20.80, 15.20, 19.73, 4.27,  0.00, 40.00),
          (0.1400, 39_780, 17.80, 19.20, 18.90, 4.10,  0.00, 40.00),
          (0.2200,183_780, 18.80, 19.20, 18.08, 3.92,  0.00, 40.00),
          (0.3300,828_000, 53.50, 21.50, 20.55, 4.45,  0.00,  0.00)]),

        ("ANEXO V — SERVICOS (tecnol., public., eng., audit. — Fator R < 28%)", "ISS",
         [(0.1550,      0, 25.00, 15.00, 14.10, 3.05, 28.85, 14.00),
          (0.1800,  4_500, 23.00, 15.00, 14.10, 3.05, 27.85, 17.00),
          (0.1950,  9_900, 24.00, 15.00, 14.92, 3.23, 23.85, 19.00),
          (0.2050, 17_100, 21.00, 15.00, 15.74, 3.41, 23.85, 21.00),
          (0.2300, 62_100, 23.00, 12.50, 14.10, 3.05, 23.85, 23.50),
          (0.3050,540_000, 35.00, 15.50, 16.44, 3.56, 29.50,  0.00)]),
    ]

    r = 1

    # ── TÍTULO ───────────────────────────────────────────────────────────
    rh(ws, r, 44)
    titulo(ws, r, "REFORMA TRIBUTARIA — SIMPLES NACIONAL — TODOS OS ANEXOS\n"
                  "CBS substitui PIS/COFINS DENTRO do DAS — aliquota nominal NAO muda | LCP 214/2025",
           span=11, sz=12)
    r += 1

    rh(ws, r, 36)
    cell(ws, r, 1,
         "CONCEITO CORRETO: A CBS entra NO LUGAR do PIS/COFINS dentro do DAS. "
         "A aliquota NOMINAL do DAS nao aumenta (ex: 4% continua 4% na 1a faixa Anexo I). "
         "A parcela que era PIS/COFINS passa a ser CBS. O PRECO DE VENDA nao muda por esse motivo.",
         bg=COR["AMARELO"], italic=True, sz=10, halign="left", span=11)
    r += 1; r += 1  # blank

    # ── CABECALHO DAS TABELAS ─────────────────────────────────────────────
    rh(ws, r, 22)
    cell(ws, r, 1, "TABELAS OFICIAIS DO SIMPLES NACIONAL — TODOS OS ANEXOS (LC 123/2006 | LC 155/2016)",
         bg=COR["AZUL_ESC"], bold=True, cor_txt="FFFFFF", sz=11, halign="center", span=11)
    r += 1; r += 1  # blank

    # ── CADA ANEXO ───────────────────────────────────────────────────────
    annex_data_start_rows = []  # para INDEX formulas na simulacao

    for ax_idx, (ax_nome, col_j_hdr, faixas) in enumerate(ANEXOS):
        rh(ws, r, 24)
        cell(ws, r, 1, ax_nome, bg=COR["AZUL_MED"], bold=True,
             cor_txt="FFFFFF", sz=11, halign="center", span=11)
        r += 1

        # cabecalho das colunas
        rh(ws, r, 32)
        hdrs = ["Fx", "RB acumulada 12m (ate R$)", "Aliq.\nNominal\nDAS",
                "Deducao\n(R$)", "IRPJ\n% partilha", "CSLL\n% partilha",
                "COFINS\n% partilha", "PIS\n% partilha", "CPP\n% partilha",
                col_j_hdr + "\n% partilha", "PIS+COFINS\n(= futura CBS)"]
        for ci, hdr in enumerate(hdrs, start=1):
            cell(ws, r, ci, hdr, bold=True, bg=COR["AZUL_CLA"], sz=9,
                 halign="center", wrap=True)
        r += 1

        # linhas de dados
        annex_data_start_rows.append(r)
        for fi, (aliq, ded, irpj, csll, cofins, pis, cpp, colj) in enumerate(faixas):
            rh(ws, r, 18)
            bg_r = COR["BRANCO"] if fi % 2 == 0 else COR["CINZA"]
            pc_total = cofins + pis

            cell(ws, r,  1, fi + 1,        bg=bg_r,           halign="center", sz=9)
            cell(ws, r,  2, LIMITES[fi],   fmt='R$ #,##0',    bg=bg_r, halign="right", sz=9)
            cell(ws, r,  3, aliq,          fmt=PCT,            bg=COR["INPUT"], halign="center", bold=True, sz=10)
            cell(ws, r,  4, ded,           fmt='R$ #,##0',    bg=bg_r, halign="right", sz=9)
            cell(ws, r,  5, irpj / 100,    fmt=PCT,            bg=bg_r, halign="center", sz=9)
            cell(ws, r,  6, csll / 100,    fmt=PCT,            bg=bg_r, halign="center", sz=9)
            cell(ws, r,  7, cofins / 100,  fmt=PCT,            bg=COR["VERM"], halign="center", sz=9)
            cell(ws, r,  8, pis / 100,     fmt=PCT,            bg=COR["VERM"], halign="center", sz=9)
            cell(ws, r,  9, cpp / 100,     fmt=PCT,            bg=bg_r, halign="center", sz=9)
            cell(ws, r, 10, colj / 100,    fmt=PCT,            bg=bg_r, halign="center", sz=9)
            cell(ws, r, 11, pc_total / 100, fmt=PCT,           bg=COR["LARANJA"], bold=True, halign="center", sz=10)
            r += 1

        # nota do anexo
        rh(ws, r, 20)
        cell(ws, r, 1,
             "Coluna PIS+COFINS (K) = parcela que sera substituida pela CBS em 2027. "
             "DAS nominal nao muda — apenas a denominacao/destinacao da parcela.",
             bg=COR["AMARELO"], italic=True, sz=9, halign="left", span=11)
        r += 1; r += 1  # blank

    # ── SIMULACAO ────────────────────────────────────────────────────────
    rh(ws, r, 26)
    cell(ws, r, 1, "SIMULACAO — IMPACTO DA CBS NO DAS: ANTES vs DEPOIS",
         bg=COR["AZUL_ESC"], bold=True, cor_txt="FFFFFF", sz=12, halign="center", span=11)
    r += 1

    rh(ws, r, 20)
    cell(ws, r, 1, "DADOS DE ENTRADA — altere os valores azuis (celulas INPUT)",
         bg=COR["AZUL_MED"], bold=True, cor_txt="FFFFFF", sz=11, halign="center", span=11)
    r += 1

    def inp_row(row, label, val, fmt, hint=""):
        rh(ws, row, 20)
        cell(ws, row, 1, label, bold=True, bg=COR["CINZA"], span=3)
        cell(ws, row, 4, val, fmt=fmt, bg=COR["INPUT"], bold=True, cor_txt="0070C0", halign="right")
        if hint:
            cell(ws, row, 5, hint, italic=True, sz=9, cor_txt="595959", span=11)

    ROW_ANX  = r
    inp_row(r, "Anexo (1=Comercio | 2=Industria | 3=ServIII | 4=ServIV | 5=ServV)", 1, "0",
            "1=Comercio  2=Industria  3=Servicos Anexo III  4=Servicos Anexo IV  5=Servicos Anexo V (Fator R<28%)")
    r += 1

    ROW_RB12 = r
    inp_row(r, "Receita Bruta acumulada 12 meses (R$)", 150_000, BRL,
            "Define a faixa do Simples Nacional — use a RB dos ultimos 12 meses")
    r += 1

    ROW_FAT  = r
    inp_row(r, "Faturamento do mes (R$)", 12_000, BRL,
            "Base de calculo do DAS mensal")
    r += 1

    ROW_CUSTO = r
    inp_row(r, "Custo do produto / servico (R$)", 7_000, BRL,
            "Custo direto para formacao do preco de venda")
    r += 1

    ROW_MARG = r
    inp_row(r, "Margem desejada (% sobre o preco de venda)", 0.20, PCT,
            "Percentual de lucro sobre o preco — ex: 20%")
    r += 1
    r += 1  # blank

    # ── CALCULO AUTOMATICO ────────────────────────────────────────────────
    rh(ws, r, 22)
    cell(ws, r, 1, "CALCULO AUTOMATICO (formulas baseadas nas tabelas acima)",
         bg=COR["VRD_ESC"], bold=True, cor_txt="FFFFFF", sz=11, halign="center", span=11)
    r += 1

    B_ANX   = f"D{ROW_ANX}"
    B_RB12  = f"D{ROW_RB12}"
    B_FAT   = f"D{ROW_FAT}"
    B_CUSTO = f"D{ROW_CUSTO}"
    B_MARG  = f"D{ROW_MARG}"

    # Faixa identificada
    ROW_FAIXA = r
    B_FAIXA   = f"D{ROW_FAIXA}"
    rh(ws, r, 20)
    cell(ws, r, 1, "Faixa identificada (1 a 6):", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4,
         f"=IF({B_RB12}<=180000,1,IF({B_RB12}<=360000,2,"
         f"IF({B_RB12}<=720000,3,IF({B_RB12}<=1800000,4,"
         f"IF({B_RB12}<=3600000,5,6)))))",
         fmt="0", bg=COR["RESULT"], bold=True, halign="center")
    cell(ws, r, 5, "Calculado automaticamente com base na RB acumulada 12m",
         italic=True, sz=9, cor_txt="595959", span=11)
    r += 1

    # Helper: CHOOSE(anexo, INDEX(col, faixa)) para cada um dos 5 anexos
    def choose_idx(col_num):
        col_letter = get_column_letter(col_num)
        parts = [f"INDEX(${col_letter}${dr}:${col_letter}${dr+5},{B_FAIXA})"
                 for dr in annex_data_start_rows]
        return f"=CHOOSE({B_ANX},{','.join(parts)})"

    # Aliquota nominal
    ROW_ALIQ_NOM = r
    B_ALIQ_NOM   = f"D{ROW_ALIQ_NOM}"
    rh(ws, r, 20)
    cell(ws, r, 1, "Aliquota nominal DAS:", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, choose_idx(3), fmt=PCT, bg=COR["RESULT"], halign="right")
    cell(ws, r, 5, "Buscada automaticamente na tabela do anexo selecionado",
         italic=True, sz=9, cor_txt="595959", span=11)
    r += 1

    # Deducao
    ROW_DED = r
    B_DED   = f"D{ROW_DED}"
    rh(ws, r, 20)
    cell(ws, r, 1, "Deducao (R$):", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, choose_idx(4), fmt=BRL, bg=COR["RESULT"], halign="right")
    cell(ws, r, 5, "R$0 = sem deducao (faixa 1 de qualquer anexo)",
         italic=True, sz=9, cor_txt="595959", span=11)
    r += 1

    # Aliquota efetiva
    ROW_EFET = r
    B_EFET   = f"D{ROW_EFET}"
    rh(ws, r, 20)
    cell(ws, r, 1, "Aliquota EFETIVA do DAS:", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, f"=({B_RB12}*{B_ALIQ_NOM}-{B_DED})/{B_RB12}",
         fmt=PCT, bg=COR["RESULT"], bold=True, halign="right")
    cell(ws, r, 5, "= (RB12 x aliq_nominal - deducao) / RB12  — formula oficial do Simples",
         italic=True, sz=9, cor_txt="595959", span=11)
    r += 1

    # % PIS+COFINS na partilha
    ROW_PCT_PC = r
    B_PCT_PC   = f"D{ROW_PCT_PC}"
    rh(ws, r, 20)
    cell(ws, r, 1, "% PIS+COFINS na partilha do DAS:", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, choose_idx(11), fmt=PCT, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, r, 5, "Porcentagem da partilha destinada a PIS+COFINS (coluna K das tabelas acima)",
         italic=True, sz=9, cor_txt="595959", span=11)
    r += 1

    # CBS embutida (formula automatica!)
    ROW_CBS = r
    B_CBS   = f"D{ROW_CBS}"
    rh(ws, r, 24)
    cell(ws, r, 1, "CBS embutida no DAS (AUTOMATICA):", bold=True, bg=COR["MARROM"],
         cor_txt="FFFFFF", sz=11, span=3)
    cell(ws, r, 4, f"={B_EFET}*{B_PCT_PC}",
         fmt=PCT, bg=COR["LARANJA"], bold=True, halign="right", sz=12)
    cell(ws, r, 5, "= aliq_efetiva x %PIS+COFINS — formula automatica baseada nas tabelas!",
         italic=True, sz=9, cor_txt="833C00", span=11)
    r += 1

    # DAS restante (sem PIS/COFINS)
    ROW_REST = r
    B_REST   = f"D{ROW_REST}"
    rh(ws, r, 20)
    cell(ws, r, 1, "DAS restante (IRPJ+CSLL+CPP+ICMS/ISS — sem PIS/COFINS):",
         bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, f"={B_EFET}-{B_CBS}",
         fmt=PCT, bg=COR["RESULT"], halign="right")
    cell(ws, r, 5, "Parte do DAS que permanece igual apos 2027 — nao e afetada pela CBS",
         italic=True, sz=9, cor_txt="595959", span=11)
    r += 1
    r += 1  # blank

    # ── FORMACAO DE PRECO ─────────────────────────────────────────────────
    rh(ws, r, 24)
    cell(ws, r, 1, "FORMACAO DE PRECO — ANTES (ate 2026) vs DEPOIS (a partir de 2027)",
         bg=COR["MARROM"], bold=True, cor_txt="FFFFFF", sz=11, halign="center", span=11)
    r += 1

    rh(ws, r, 30)
    for ci, hdr in enumerate(["Item", "", "", "ANTES (2026)", "DEPOIS (2027+)", "Variacao (R$)", "Observacao"],
                              start=1):
        cell(ws, r, ci, hdr, bold=True, bg=COR["AZUL_CLA"], halign="center", sz=9, wrap=True)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=11)
    r += 1

    # Preco = Custo / (1 - DAS_efetiva - Margem)
    # Formula IDENTICA antes e depois: DAS nominal nao muda!
    ROW_PRECO = r
    F_PRECO = f"={B_CUSTO}/(1-{B_EFET}-{B_MARG})"
    rh(ws, r, 22)
    cell(ws, r, 1, "Preco de venda (R$)", bold=True, span=3)
    cell(ws, r, 4, F_PRECO, fmt=BRL, bg=COR["AMARELO"], bold=True, halign="right")
    cell(ws, r, 5, F_PRECO, fmt=BRL, bg=COR["AMARELO"], bold=True, halign="right")
    cell(ws, r, 6, f"=E{r}-D{r}", fmt=BRL, bg=COR["CINZA"], halign="right")
    cell(ws, r, 7, "Identico: DAS nominal nao muda (0% variacao)", italic=True, sz=9, span=11)
    r += 1

    ROW_DAS_VAL = r
    rh(ws, r, 20)
    cell(ws, r, 1, "DAS total pago (R$)", span=3)
    cell(ws, r, 4, f"=D{ROW_PRECO}*{B_EFET}", fmt=BRL, bg=COR["LARANJA"], halign="right")
    cell(ws, r, 5, f"=E{ROW_PRECO}*{B_EFET}", fmt=BRL, bg=COR["LARANJA"], halign="right")
    cell(ws, r, 6, f"=E{r}-D{r}", fmt=BRL, bg=COR["CINZA"], halign="right")
    cell(ws, r, 7, "Identico: carga total do DAS igual (0%)", italic=True, sz=9, span=11)
    r += 1

    ROW_PC_VAL = r
    rh(ws, r, 18)
    cell(ws, r, 1, "  -> PIS+COFINS dentro do DAS (ANTES)", italic=True, sz=9, span=3)
    cell(ws, r, 4, f"=D{ROW_PRECO}*{B_CBS}", fmt=BRL, bg=COR["VERM"], halign="right", sz=9)
    cell(ws, r, 5, "---", halign="center", bg=COR["CINZA_ESC"], sz=9)
    cell(ws, r, 6, "", bg=COR["CINZA"])
    cell(ws, r, 7, "Extinta em 2027 — substituida pela CBS", italic=True, sz=9,
         cor_txt="CC0000", span=11)
    r += 1

    ROW_CBS_VAL = r
    rh(ws, r, 18)
    cell(ws, r, 1, "  -> CBS embutida no DAS (DEPOIS — formula automatica)", italic=True, sz=9, span=3)
    cell(ws, r, 4, "---", halign="center", bg=COR["CINZA_ESC"], sz=9)
    cell(ws, r, 5, f"=E{ROW_PRECO}*{B_CBS}", fmt=BRL, bg=COR["LARANJA"], halign="right", sz=9)
    cell(ws, r, 6, f"=E{ROW_CBS_VAL}*{B_CBS}-D{ROW_PC_VAL}", fmt=BRL, bg=COR["CINZA"], halign="right")
    cell(ws, r, 7, "CBS substitui PIS/COFINS dentro do DAS — mesmo valor!", italic=True, sz=9,
         cor_txt="994C00", span=11)
    r += 1

    ROW_REST_VAL = r
    rh(ws, r, 18)
    cell(ws, r, 1, "  -> Demais tributos DAS (IRPJ+CSLL+CPP+ICMS/ISS)", italic=True, sz=9, span=3)
    cell(ws, r, 4, f"=D{ROW_PRECO}*{B_REST}", fmt=BRL, halign="right", sz=9)
    cell(ws, r, 5, f"=E{ROW_PRECO}*{B_REST}", fmt=BRL, halign="right", sz=9)
    cell(ws, r, 6, f"=E{r}-D{r}", fmt=BRL, bg=COR["CINZA"], halign="right")
    cell(ws, r, 7, "Permanece inalterado", italic=True, sz=9, span=11)
    r += 1

    ROW_LUCRO = r
    rh(ws, r, 22)
    cell(ws, r, 1, "Lucro bruto (R$)", bold=True, span=3)
    cell(ws, r, 4, f"=D{ROW_PRECO}-D{ROW_DAS_VAL}-{B_CUSTO}",
         fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, r, 5, f"=E{ROW_PRECO}-E{ROW_DAS_VAL}-{B_CUSTO}",
         fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, r, 6, f"=E{r}-D{r}", fmt=BRL, bg=COR["CINZA"], halign="right")
    cell(ws, r, 7, "Identico: nenhum impacto no lucro", italic=True, sz=9, span=11)
    r += 1
    r += 1  # blank

    # ── IMPACTO ───────────────────────────────────────────────────────────
    rh(ws, r, 22)
    cell(ws, r, 1, "IMPACTO RESUMIDO DA REFORMA NO SIMPLES NACIONAL",
         bg=COR["AZUL_MED"], bold=True, cor_txt="FFFFFF", sz=11, halign="center", span=11)
    r += 1

    impactos = [
        ("Variacao no preco ao consumidor",
         f"=E{ROW_PRECO}-D{ROW_PRECO}", BRL,
         f"=(E{ROW_PRECO}-D{ROW_PRECO})/D{ROW_PRECO}", PCT,
         "0%: aliquota nominal DAS nao muda = preco igual"),
        ("Variacao na carga tributaria (DAS)",
         f"=E{ROW_DAS_VAL}-D{ROW_DAS_VAL}", BRL,
         f"=(E{ROW_DAS_VAL}-D{ROW_DAS_VAL})/D{ROW_DAS_VAL}", PCT,
         "0%: DAS total permanece o mesmo"),
        ("CBS visivel embutida no DAS (valor mensal)",
         f"=E{ROW_PRECO}*{B_CBS}", BRL,
         f"={B_CBS}", PCT,
         "Novo: CBS e identificada e pode gerar credito CBS ao cliente na cadeia"),
    ]

    for lbl, v1, f1, v2, f2, obs in impactos:
        rh(ws, r, 20)
        cell(ws, r, 1, lbl, bold=True, span=3)
        cell(ws, r, 4, v1, fmt=f1, bg=COR["RESULT"], halign="right", bold=True)
        cell(ws, r, 5, v2, fmt=f2, bg=COR["RESULT"], halign="right")
        cell(ws, r, 6, obs, italic=True, sz=9, span=11)
        r += 1

    r += 1  # blank

    # ── NOTAS FINAIS ──────────────────────────────────────────────────────
    notas_finais = [
        "CONCLUSAO: A CBS NAO e cobrada por fora no Simples Nacional. Ela entra NO LUGAR do PIS/COFINS dentro do DAS.",
        "PRECO: Como a aliquota nominal nao muda, o preco pela formula Custo/(1-DAS_efetiva-Margem) e IDENTICO antes e depois.",
        "VISIBILIDADE: A mudanca real e que a CBS torna-se identificada — potencialmente gerando credito de CBS para o cliente.",
        "FAIXA 6 (RB12 > R$3,6M): ICMS/ISS cai a 0% no DAS — empresa paga esses tributos separadamente. Verifique impacto.",
        "FATOR R (Anexo V): Se folha/receita >= 28%, a empresa migra para o Anexo III — aliquotas menores. Confira seu caso.",
        "ANEXO IV (sem CPP): CPP e paga fora do DAS — considere no custo total da folha de pagamento.",
        "FONTE: LC 123/2006, LC 155/2016, LC 167/2019 (tabelas de partilha) | LCP 214/2025 (CBS substituindo PIS/COFINS no DAS).",
    ]

    for n in notas_finais:
        rh(ws, r, 20)
        c = ws.cell(row=r, column=1, value=f"* {n}")
        c.font = Font(name="Calibri", italic=True, size=9, color="595959")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill = PatternFill(start_color=COR["AMARELO"], end_color=COR["AMARELO"], fill_type="solid")
        c.border = bd()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        r += 1


# ═══════════════════════════════════════════════════════════════════════
#   SHEET 4 — COMPARATIVO GERAL
# ═══════════════════════════════════════════════════════════════════════
def sheet_comparativo(wb):
    ws = wb.create_sheet("Comparativo Geral")
    for col, w in zip("ABCDEFG", [30, 18, 18, 18, 18, 18, 22]):
        cw(ws, col, w)

    rh(ws, 1, 48)
    titulo(ws, 1, "COMPARATIVO GERAL — TODOS OS REGIMES\n"
                  "Reforma Tributária Brasileira | LCP 214/2025 | CBS em 2027", sz=13, span=7)
    rh(ws, 2, 20)
    cell(ws, 2, 1, "Resumo side-by-side dos três regimes. Os detalhes ficam nas abas individuais de cada regime.",
         bg=COR["AZUL_CLA"], italic=True, sz=10, halign="center", span=7)
    rh(ws, 3, 8)

    # Cabeçalhos
    rh(ws, 4, 35)
    cell(ws, 4, 1, "CARACTERÍSTICA", bg=COR["AZUL_ESC"], bold=True, cor_txt="FFFFFF", halign="center")
    cell(ws, 4, 2, "LUCRO PRESUMIDO\nANTES", bg=COR["VRD_ESC"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 4, 3, "LUCRO PRESUMIDO\nDEPOIS", bg=COR["MARROM"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 4, 4, "LUCRO REAL\nANTES", bg=COR["VRD_ESC"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 4, 5, "LUCRO REAL\nDEPOIS", bg=COR["MARROM"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 4, 6, "SIMPLES NACIONAL\nANTES", bg=COR["VRD_ESC"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)
    cell(ws, 4, 7, "SIMPLES NACIONAL\nDEPOIS (estimado)", bg=COR["MARROM"], bold=True, cor_txt="FFFFFF", halign="center", sz=10)

    linhas = [
        ("Tributo sobre receita/venda", "PIS+COFINS", "CBS", "PIS+COFINS", "CBS", "PIS+COFINS (no DAS)", "CBS favorecida"),
        ("Alíquota nominal", "3,65%", "~8,8%*", "9,25%", "~8,8%*", "Varia por faixa", "~4,4%* (estimado)"),
        ("Forma de cálculo", "POR DENTRO\n(embutido no preço)", "POR FORA\n(destacado)", "POR DENTRO\n(embutido no preço)", "POR FORA\n(destacado)", "% sobre receita\n(no DAS)", "% sobre receita\n(por fora)"),
        ("Crédito sobre insumos", "NÃO", "SIM ✓", "SIM ✓", "SIM ✓", "NÃO", "SIM ✓ (limitado)"),
        ("Recolhimento", "Empresa recolhe\nmanualmente (DARF)", "Split Payment\nautomático", "Empresa recolhe\nmanualmente (DARF)", "Split Payment\nautomático", "DAS unificado\n(boleto mensal)", "DAS + CBS separado"),
        ("Impacto na precificação", "PIS+COFINS embutidos\nno preço", "Retirar PIS+COFINS;\nCBS por cima", "PIS+COFINS embutidos\nno preço", "Retirar PIS+COFINS;\nCBS por cima", "Tributo invisível\nno preço", "CBS destacada\nao consumidor"),
        ("Risco de carga maior", "Benchmark", "Menor se tiver\nmuitos insumos", "Benchmark", "Similar ou menor", "Benchmark", "Depende da\nalíquota final"),
        ("Empresas mais beneficiadas", "—", "Muitos insumos\n(indústria/comércio)", "—", "Todas (já há crédito)", "—", "Quem tem poucos\ncustos hoje"),
        ("Empresas mais afetadas", "—", "Poucos insumos\n(serviços puros)", "—", "Impacto mínimo", "—", "Serviços com poucas\ndespesas dedutíveis"),
        ("Vigência", "Até 31/12/2026", "A partir de 01/01/2027", "Até 31/12/2026", "A partir de 01/01/2027", "Até 31/12/2026", "A partir de 01/01/2027"),
    ]

    bgs_alt = [COR["CINZA"], COR["BRANCO"]]
    bgs_dep = [COR["AMARELO"], "FFF5E0"]
    for i, linha in enumerate(linhas):
        r = 5 + i
        rh(ws, r, 40)
        bg_a = bgs_alt[i % 2]
        bg_d = bgs_dep[i % 2]
        cell(ws, r, 1, linha[0], bg=COR["CINZA_ESC"], bold=True, sz=10)
        cell(ws, r, 2, linha[1], bg=bg_a, halign="center", sz=10)
        cell(ws, r, 3, linha[2], bg=bg_d, halign="center", sz=10)
        cell(ws, r, 4, linha[3], bg=bg_a, halign="center", sz=10)
        cell(ws, r, 5, linha[4], bg=bg_d, halign="center", sz=10)
        cell(ws, r, 6, linha[5], bg=bg_a, halign="center", sz=10)
        cell(ws, r, 7, linha[6], bg=bg_d, halign="center", sz=10)

    rh(ws, 16, 8)

    # ── LINHA DO TEMPO
    rh(ws, 17, 28)
    cell(ws, 17, 1, "LINHA DO TEMPO DA TRANSIÇÃO", bg=COR["AZUL_ESC"], bold=True, cor_txt="FFFFFF",
         sz=12, halign="center", span=7)

    timeline = [
        ("2026", "Ano-teste: CBS 0,9% + IBS 0,1% coexistindo com PIS/COFINS/ICMS/ISS. Sem substituição ainda.", COR["AZUL_CLA"]),
        ("2027", "CBS substitui integralmente PIS e COFINS. Todas as empresas migram. IBS ainda pequeno.", COR["VRD_CLA"]),
        ("2029", "IBS começa a substituir ICMS e ISS gradualmente. Período de coexistência.", COR["AMARELO"]),
        ("2029–2032", "Redução progressiva de ICMS e ISS. Aumento progressivo do IBS. Transição gradual por 4 anos.", COR["LARANJA"]),
        ("2033+", "Sistema completo: apenas IBS + CBS. ICMS, ISS, PIS e COFINS extintos definitivamente.", COR["VRD_CLA"]),
    ]
    for i, (ano, desc, bg) in enumerate(timeline):
        r = 18 + i
        rh(ws, r, 30)
        cell(ws, r, 1, ano, bg=COR["AZUL_ESC"], bold=True, cor_txt="FFFFFF", halign="center", sz=12)
        cell(ws, r, 2, desc, bg=bg, sz=10, span=7)

    rh(ws, 24, 8)
    rh(ws, 25, 22)
    secao(ws, 25, "LEGENDA E AVISOS", bg="MARROM")
    avisos = [
        "* Alíquotas CBS são ESTIMATIVAS. Serão fixadas por Resolução do Senado Federal (Art. 18, LCP 214). Esta planilha usa parâmetros de mercado para simulação.",
        "ANTES: PIS/COFINS por dentro = o preço JÁ INCLUI o tributo. O consumidor paga sem ver separado (exceto na NF).",
        "DEPOIS: CBS por fora = o preço base NÃO inclui o tributo. O CBS é somado ao preço e destacado. Consumer vê separado.",
        "SPLIT PAYMENT (Arts. 31-34): O próprio sistema financeiro (banco, maquininha, Pix) recolhe automaticamente a CBS.",
        "ART. 12 §2° V da LCP 214: Durante 2026-2032, ICMS, ISS, PIS e COFINS NÃO entram na base do CBS/IBS. Sem dupla tributação.",
        "RECOMENDAÇÃO: Revise sua precificação em 2026 antes da migração. Consulte seu contador para análise específica do seu negócio.",
    ]
    for i, av in enumerate(avisos):
        nota(ws, 26 + i, av)


# ═══════════════════════════════════════════════════════════════════════
#   SHEET 5 — REGIME HIBRIDO SIMPLES NACIONAL (ICMS-ST)
# ═══════════════════════════════════════════════════════════════════════
def sheet_regime_hibrido(wb):
    ws = wb.create_sheet("Regime Hibrido SN")
    for col, w in zip("ABCDEFG", [4, 28, 18, 18, 18, 18, 20]):
        cw(ws, col, w)

    # Rows onde os dados de cada anexo começam na aba 'Simples Nacional'
    ANX_ROWS = [8, 18, 28, 38, 48]

    r = 1
    rh(ws, r, 48)
    titulo(ws, r,
           "REGIME HIBRIDO — SIMPLES NACIONAL COM ICMS-ST\n"
           "Empresa optante do SN com parte da receita sujeita a Substituicao Tributaria | LCP 214/2025",
           sz=12, span=7)
    r += 1

    rh(ws, r, 36)
    cell(ws, r, 1,
         "ICMS-ST (Substituicao Tributaria): o ICMS e recolhido pelo fabricante/distribuidor antes de chegar a voce. "
         "Por isso, a parcela de ICMS que seria do DAS ja foi paga 'por fora' — e o DAS e reduzido nessa parte. "
         "Esta aba permite simular o impacto para empresas com mix de receitas (parte normal, parte com ST).",
         bg=COR["AMARELO"], italic=True, sz=10, span=7)
    r += 1; r += 1

    # ── INPUTS ────────────────────────────────────────────────────────────
    rh(ws, r, 22)
    secao(ws, r, "DADOS DE ENTRADA — altere as celulas azuis", span=7)
    r += 1

    ROW_ANX  = r
    rh(ws, r, 20)
    cell(ws, r, 1, "Anexo do Simples Nacional (1=Comerc | 2=Ind | 3=SrvIII | 4=SrvIV | 5=SrvV)",
         bg=COR["CINZA"], bold=True, span=3)
    cell(ws, r, 4, 1, fmt="0", bg=COR["INPUT"], bold=True, cor_txt="0070C0", halign="center")
    cell(ws, r, 5, "1=Comercio  2=Industria  3=Servicos III  4=Servicos IV  5=Servicos V (Fator R<28%)",
         italic=True, sz=9, cor_txt="595959", span=7)
    r += 1

    ROW_RB12 = r
    rh(ws, r, 20)
    cell(ws, r, 1, "Receita Bruta acumulada 12 meses (R$)", bg=COR["CINZA"], bold=True, span=3)
    cell(ws, r, 4, 150_000, fmt=BRL, bg=COR["INPUT"], bold=True, cor_txt="0070C0", halign="right")
    cell(ws, r, 5, "Define a faixa — use os 12 meses anteriores", italic=True, sz=9, cor_txt="595959", span=7)
    r += 1

    ROW_PCT_ST = r
    rh(ws, r, 20)
    cell(ws, r, 1, "% da receita mensal sujeita a ICMS-ST", bg=COR["CINZA"], bold=True, span=3)
    cell(ws, r, 4, 0.50, fmt=PCT, bg=COR["INPUT"], bold=True, cor_txt="0070C0", halign="right")
    cell(ws, r, 5, "Ex: 50% = metade das vendas tem ICMS recolhido pelo fornecedor via ST", italic=True, sz=9, cor_txt="595959", span=7)
    r += 1

    ROW_CUSTO = r
    rh(ws, r, 20)
    cell(ws, r, 1, "Custo do produto (R$)", bg=COR["CINZA"], bold=True, span=3)
    cell(ws, r, 4, 1_000, fmt=BRL, bg=COR["INPUT"], bold=True, cor_txt="0070C0", halign="right")
    cell(ws, r, 5, "Custo de aquisicao — base para formacao do preco de venda", italic=True, sz=9, cor_txt="595959", span=7)
    r += 1

    ROW_MARG = r
    rh(ws, r, 20)
    cell(ws, r, 1, "Margem desejada (% sobre o preco de venda)", bg=COR["CINZA"], bold=True, span=3)
    cell(ws, r, 4, 0.20, fmt=PCT, bg=COR["INPUT"], bold=True, cor_txt="0070C0", halign="right")
    cell(ws, r, 5, "Ex: 20% de margem sobre o preco de venda", italic=True, sz=9, cor_txt="595959", span=7)
    r += 1; r += 1

    # ── CALCULO AUTOMATICO ────────────────────────────────────────────────
    rh(ws, r, 22)
    secao(ws, r, "CALCULO AUTOMATICO — baseado nas tabelas do Simples Nacional", span=7)
    r += 1

    B_ANX    = f"D{ROW_ANX}"
    B_RB12   = f"D{ROW_RB12}"
    B_PCT_ST = f"D{ROW_PCT_ST}"
    B_CUSTO  = f"D{ROW_CUSTO}"
    B_MARG   = f"D{ROW_MARG}"

    # Substitui placeholder D_ANX e D_FAIXA depois de criar as variaveis
    def csn(col_letter, b_faixa_ref):
        parts = [f"INDEX('Simples Nacional'!${col_letter}${dr}:${col_letter}${dr+5},{b_faixa_ref})"
                 for dr in ANX_ROWS]
        return f"=CHOOSE({B_ANX},{','.join(parts)})"

    # Faixa
    ROW_FAIXA = r
    B_FAIXA   = f"D{ROW_FAIXA}"
    rh(ws, r, 20)
    cell(ws, r, 1, "Faixa identificada:", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4,
         f"=IF({B_RB12}<=180000,1,IF({B_RB12}<=360000,2,"
         f"IF({B_RB12}<=720000,3,IF({B_RB12}<=1800000,4,"
         f"IF({B_RB12}<=3600000,5,6)))))",
         fmt="0", bg=COR["RESULT"], bold=True, halign="center")
    cell(ws, r, 5, "Calculado pela RB12", italic=True, sz=9, cor_txt="595959", span=7)
    r += 1

    # Aliq nominal
    ROW_NOM = r
    B_NOM   = f"D{ROW_NOM}"
    rh(ws, r, 20)
    cell(ws, r, 1, "Aliquota nominal DAS:", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, csn("C", B_FAIXA), fmt=PCT, bg=COR["RESULT"], halign="right")
    cell(ws, r, 5, "Buscada automaticamente na tabela do Simples Nacional", italic=True, sz=9, cor_txt="595959", span=7)
    r += 1

    # Deducao
    ROW_DED = r
    B_DED   = f"D{ROW_DED}"
    rh(ws, r, 20)
    cell(ws, r, 1, "Deducao (R$):", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, csn("D", B_FAIXA), fmt=BRL, bg=COR["RESULT"], halign="right")
    r += 1

    # Aliq efetiva
    ROW_EFET = r
    B_EFET   = f"D{ROW_EFET}"
    rh(ws, r, 20)
    cell(ws, r, 1, "Aliquota EFETIVA do DAS:", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, f"=({B_RB12}*{B_NOM}-{B_DED})/{B_RB12}", fmt=PCT, bg=COR["RESULT"], bold=True, halign="right")
    cell(ws, r, 5, "= (RB12 x aliq_nom - deducao) / RB12", italic=True, sz=9, cor_txt="595959", span=7)
    r += 1

    # % ICMS na partilha (col J = coluna 10)
    ROW_ICMS_P = r
    B_ICMS_P   = f"D{ROW_ICMS_P}"
    rh(ws, r, 20)
    cell(ws, r, 1, "% ICMS/ISS na partilha do DAS (col J):", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, csn("J", B_FAIXA), fmt=PCT, bg=COR["LARANJA"], halign="right")
    cell(ws, r, 5, "Esta e a parcela que o fornecedor ja recolheu via ST — sera zerada no DAS", italic=True, sz=9, cor_txt="595959", span=7)
    r += 1

    # % PIS+COFINS (col K = coluna 11)
    ROW_PC_P = r
    B_PC_P   = f"D{ROW_PC_P}"
    rh(ws, r, 20)
    cell(ws, r, 1, "% PIS+COFINS / CBS embutida na partilha (col K):", bold=True, bg=COR["CINZA"], span=3)
    cell(ws, r, 4, csn("K", B_FAIXA), fmt=PCT, bg=COR["LARANJA"], halign="right")
    cell(ws, r, 5, "Parcela que sera identificada como CBS a partir de 2027", italic=True, sz=9, cor_txt="595959", span=7)
    r += 1

    # DAS com ST (ICMS zerado)
    ROW_DAS_ST = r
    B_DAS_ST   = f"D{ROW_DAS_ST}"
    rh(ws, r, 20)
    cell(ws, r, 1, "Aliq efetiva DAS c/ ICMS-ST (ICMS = 0 no DAS):", bold=True, bg=COR["MARROM"], cor_txt="FFFFFF", span=3)
    cell(ws, r, 4, f"={B_EFET}*(1-{B_ICMS_P})", fmt=PCT, bg=COR["LARANJA"], bold=True, halign="right")
    cell(ws, r, 5, "= aliq_efet x (1 - % ICMS): ICMS ja pago pelo fornecedor via ST", italic=True, sz=9, cor_txt="833C00", span=7)
    r += 1; r += 1

    # ── SIMULACAO DE PRECO ────────────────────────────────────────────────
    rh(ws, r, 28)
    cell(ws, r, 1, "SIMULACAO DE PRECO E RESULTADO",
         bg=COR["AZUL_ESC"], bold=True, cor_txt="FFFFFF", sz=11, halign="center", span=7)
    r += 1

    rh(ws, r, 32)
    for ci, hdr in enumerate(
        ["Item", "", "", "SN NORMAL\n(sem ST)", "SN COM ICMS-ST\n(ICMS=0 no DAS)",
         "HIBRIDO PONDERADO\n(mix normal + ST)", "Observacao"], start=1):
        cell(ws, r, ci, hdr, bold=True, bg=COR["AZUL_MED"], cor_txt="FFFFFF", halign="center", sz=9)
    r += 1

    # Preco de venda
    ROW_PRECO = r
    rh(ws, r, 22)
    cell(ws, r, 1, "Preco de venda (R$)", bold=True, span=3)
    F_PRECO_NORM = f"={B_CUSTO}/(1-{B_EFET}-{B_MARG})"
    F_PRECO_ST   = f"={B_CUSTO}/(1-{B_DAS_ST}-{B_MARG})"
    F_PRECO_HYB  = f"=(1-{B_PCT_ST})*D{ROW_PRECO}+{B_PCT_ST}*E{ROW_PRECO}"
    cell(ws, r, 4, F_PRECO_NORM, fmt=BRL, bg=COR["AMARELO"], bold=True, halign="right")
    cell(ws, r, 5, F_PRECO_ST,   fmt=BRL, bg=COR["AMARELO"], bold=True, halign="right")
    cell(ws, r, 6, F_PRECO_HYB,  fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, r, 7, "ST reduz o DAS → preco de venda menor", italic=True, sz=9, span=7)
    r += 1

    # DAS pago
    ROW_DAS = r
    rh(ws, r, 20)
    cell(ws, r, 1, "DAS efetivo pago (R$)", span=3)
    cell(ws, r, 4, f"=D{ROW_PRECO}*{B_EFET}",   fmt=BRL, bg=COR["LARANJA"], halign="right")
    cell(ws, r, 5, f"=E{ROW_PRECO}*{B_DAS_ST}",  fmt=BRL, bg=COR["LARANJA"], halign="right")
    cell(ws, r, 6, f"=(1-{B_PCT_ST})*D{ROW_DAS}+{B_PCT_ST}*E{ROW_DAS}",
         fmt=BRL, bg=COR["LARANJA"], halign="right")
    cell(ws, r, 7, "DAS reduzido nas vendas com ST", italic=True, sz=9, span=7)
    r += 1

    # CBS embutida no DAS
    ROW_CBS = r
    rh(ws, r, 18)
    cell(ws, r, 1, "  -> CBS embutida no DAS (% efetivo x % PC)", italic=True, sz=9, span=3)
    cell(ws, r, 4, f"=D{ROW_PRECO}*{B_EFET}*{B_PC_P}", fmt=BRL, bg=COR["LARANJA"], halign="right", sz=9)
    cell(ws, r, 5, f"=E{ROW_PRECO}*{B_DAS_ST}*{B_PC_P}", fmt=BRL, bg=COR["LARANJA"], halign="right", sz=9)
    cell(ws, r, 6, f"=(1-{B_PCT_ST})*D{ROW_CBS}+{B_PCT_ST}*E{ROW_CBS}",
         fmt=BRL, bg=COR["LARANJA"], halign="right", sz=9)
    cell(ws, r, 7, "CBS identificada dentro do DAS (sera visivel apos 2027)", italic=True, sz=9, span=7)
    r += 1

    # Lucro
    ROW_LUC = r
    rh(ws, r, 22)
    cell(ws, r, 1, "Lucro bruto (R$)", bold=True, span=3)
    cell(ws, r, 4, f"=D{ROW_PRECO}-D{ROW_DAS}-{B_CUSTO}", fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, r, 5, f"=E{ROW_PRECO}-E{ROW_DAS}-{B_CUSTO}", fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, r, 6, f"=F{ROW_PRECO}-F{ROW_DAS}-{B_CUSTO}", fmt=BRL, bg=COR["VRD_CLA"], bold=True, halign="right")
    cell(ws, r, 7, "= Preco - DAS - Custo", italic=True, sz=9, span=7)
    r += 1

    # Margem efetiva
    rh(ws, r, 18)
    cell(ws, r, 1, "Margem efetiva (%)", span=3)
    cell(ws, r, 4, f"=D{ROW_LUC}/D{ROW_PRECO}", fmt=PCT, bg=COR["CINZA"], halign="right")
    cell(ws, r, 5, f"=E{ROW_LUC}/E{ROW_PRECO}", fmt=PCT, bg=COR["CINZA"], halign="right")
    cell(ws, r, 6, f"=F{ROW_LUC}/F{ROW_PRECO}", fmt=PCT, bg=COR["CINZA"], halign="right")
    r += 1

    # Carga efetiva %
    rh(ws, r, 18)
    cell(ws, r, 1, "Carga DAS efetiva (% do preco)", span=3)
    cell(ws, r, 4, f"=D{ROW_DAS}/D{ROW_PRECO}", fmt=PCT, bg=COR["CINZA"], halign="right")
    cell(ws, r, 5, f"=E{ROW_DAS}/E{ROW_PRECO}", fmt=PCT, bg=COR["CINZA"], halign="right")
    cell(ws, r, 6, f"=F{ROW_DAS}/F{ROW_PRECO}", fmt=PCT, bg=COR["CINZA"], halign="right")
    r += 1; r += 1

    # ── IMPACTO ST ────────────────────────────────────────────────────────
    rh(ws, r, 22)
    secao(ws, r, "IMPACTO DO ICMS-ST — NORMAL vs HIBRIDO", span=7)
    r += 1

    impactos_st = [
        ("Reducao no DAS (R$) — ST vs Normal",
         f"=E{ROW_DAS}-D{ROW_DAS}", BRL,
         f"=(E{ROW_DAS}-D{ROW_DAS})/D{ROW_DAS}", PCT,
         "Negativo = reducao do DAS (beneficio da ST)"),
        ("Reducao no preco de venda (R$) — Hibrido vs Normal",
         f"=F{ROW_PRECO}-D{ROW_PRECO}", BRL,
         f"=(F{ROW_PRECO}-D{ROW_PRECO})/D{ROW_PRECO}", PCT,
         "Preco hibrido ponderado vs preco integral sem ST"),
        ("Ganho no lucro (R$) — ST vs Normal",
         f"=E{ROW_LUC}-D{ROW_LUC}", BRL,
         f"=(E{ROW_LUC}-D{ROW_LUC})/D{ROW_LUC}", PCT,
         "ST reduz DAS → margem maior com mesmo preco"),
    ]

    rh(ws, r, 28)
    for ci, hdr in enumerate(["Indicador", "", "", "Variacao (R$)", "Variacao (%)", "", "Observacao"], start=1):
        cell(ws, r, ci, hdr, bold=True, bg=COR["AZUL_CLA"], halign="center", sz=9)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    r += 1

    for lbl, v1, f1, v2, f2, obs in impactos_st:
        rh(ws, r, 22)
        cell(ws, r, 1, lbl, bold=True, span=3)
        cell(ws, r, 4, v1, fmt=f1, bg=COR["RESULT"], halign="right", bold=True)
        cell(ws, r, 5, v2, fmt=f2, bg=COR["RESULT"], halign="right")
        cell(ws, r, 6, "", bg=COR["CINZA"])
        cell(ws, r, 7, obs, italic=True, sz=9)
        r += 1

    r += 1
    notas_h = [
        "ICMS-ST: O ICMS e recolhido 'por substituicao' pelo fabricante/importador antes de chegar ao varejista optante do SN.",
        "EFEITO NO DAS: A parcela de ICMS da partilha (col J das tabelas) e zerada para o varejista — DAS menor.",
        "PRECO COMPETITIVO: Com DAS menor, o varejista com ST pode praticar preco menor mantendo a mesma margem.",
        "CUIDADO: Nem todos os produtos tem ST em todos os estados — confirme a lista de NCMs sujeitos a ST no seu estado.",
        "PLANEJAMENTO: Se voce compra mercadorias com ST, inclua isso na precificacao e compare o DAS efetivo ao DAS normal.",
    ]
    for n in notas_h:
        rh(ws, r, 18)
        c = ws.cell(row=r, column=1, value=f"* {n}")
        c.font = Font(name="Calibri", italic=True, size=9, color="595959")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.fill = PatternFill(start_color=COR["AMARELO"], end_color=COR["AMARELO"], fill_type="solid")
        c.border = bd()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 1


# ═══════════════════════════════════════════════════════════════════════
#   SHEET 6 — RELATORIO CLIENTE (substitui Simulacao Comparativa)
# ═══════════════════════════════════════════════════════════════════════
def sheet_relatorio_cliente(wb):
    ws = wb.create_sheet("Relatorio Cliente")
    for col, w in zip("ABCDE", [32, 22, 22, 22, 26]):
        cw(ws, col, w)

    LP = "'Lucro Presumido'"
    LR = "'Lucro Real'"
    SN = "'Simples Nacional'"

    # ── CABECALHO / IDENTIFICACAO ─────────────────────────────────────────
    rh(ws, 1, 52)
    titulo(ws, 1,
           "RELATORIO DE PLANEJAMENTO TRIBUTARIO — REFORMA TRIBUTARIA\n"
           "Impacto da substituicao de PIS/COFINS pela CBS | LCP 214/2025",
           sz=13, span=5)

    rh(ws, 2, 8)

    rh(ws, 3, 22)
    secao(ws, 3, "IDENTIFICACAO DO CLIENTE E DO RELATORIO", span=5)

    dados_emp = [
        (4,  "Empresa / Cliente:",     "",             "Preencha o nome da empresa"),
        (5,  "CNPJ:",                  "",             "CNPJ do cliente"),
        (6,  "Atividade / Ramo:",      "",             "Descricao da atividade economica"),
        (7,  "Regime Tributario Atual:","",            "Lucro Presumido / Lucro Real / Simples Nacional"),
        (8,  "Consultor Responsavel:", "",             "Nome do consultor"),
        (9,  "Data do Relatorio:",     "=TODAY()",     "Data de emissao (automatica)"),
    ]
    for row, lbl, val, hint in dados_emp:
        rh(ws, row, 20)
        cell(ws, row, 1, lbl, bg=COR["CINZA"], bold=True)
        cell(ws, row, 2, val, fmt="DD/MM/YYYY" if row == 9 else None,
             bg=COR["INPUT"], bold=True, cor_txt="0070C0", halign="left", span=3)
        cell(ws, row, 5, hint, italic=True, sz=9, cor_txt="595959")

    rh(ws, 10, 8)

    # ── SECAO 1: PARAMETROS UTILIZADOS ───────────────────────────────────
    rh(ws, 11, 22)
    secao(ws, 11, "1. PARAMETROS UTILIZADOS NAS SIMULACOES", span=5)

    rh(ws, 12, 32)
    for ci, hdr in enumerate(["Parametro", "Lucro Presumido", "Lucro Real", "Simples Nacional", "Obs"], start=1):
        cell(ws, 12, ci, hdr, bold=True, bg=COR["AZUL_ESC"], cor_txt="FFFFFF", halign="center", sz=10)

    params = [
        ("Custo de aquisicao / producao",
         f"={LP}!B5", f"={LR}!B5", f"={SN}!D61", BRL, "Base para todos os calculos"),
        ("Margem desejada (% sobre preco)",
         f"={LP}!B6", f"={LR}!B6", f"={SN}!D62", PCT, ""),
        ("Aliquota CBS / DAS efetiva",
         f"={LP}!B7", f"={LR}!B7", f"={SN}!D68", PCT, "LP/LR = CBS por fora; SN = DAS efetivo"),
        ("% credito sobre insumos",
         f"={LP}!B8", f"={LR}!B8", "n/a", PCT, "SN nao tem credito (DAS unificado)"),
    ]
    for ri, (lbl, v_lp, v_lr, v_sn, fmt, obs) in enumerate(params, start=13):
        rh(ws, ri, 20)
        bg = COR["CINZA"] if ri % 2 == 0 else COR["BRANCO"]
        cell(ws, ri, 1, lbl, bg=bg, bold=True)
        cell(ws, ri, 2, v_lp, fmt=fmt if v_lp != "n/a" else None, bg=bg, halign="right")
        cell(ws, ri, 3, v_lr, fmt=fmt if v_lr != "n/a" else None, bg=bg, halign="right")
        if v_sn == "n/a":
            cell(ws, ri, 4, "n/a", bg=bg, halign="center", italic=True, cor_txt="AAAAAA")
        else:
            cell(ws, ri, 4, v_sn, fmt=fmt, bg=bg, halign="right")
        cell(ws, ri, 5, obs, italic=True, sz=9, cor_txt="595959")

    rh(ws, 17, 8)

    # ── SECAO 2: SITUACAO ATUAL (ANTES DA REFORMA) ────────────────────────
    rh(ws, 18, 22)
    secao(ws, 18, "2. SITUACAO ATUAL — ANTES DA REFORMA (ate 2026)", span=5)

    rh(ws, 19, 32)
    for ci, hdr in enumerate(["Indicador", "Lucro Presumido", "Lucro Real", "Simples Nacional", ""], start=1):
        cell(ws, 19, ci, hdr, bold=True, bg=COR["VRD_ESC"], cor_txt="FFFFFF", halign="center", sz=10)

    antes_rows = [
        ("Preco de venda / base (R$)",         f"={LP}!B12", f"={LR}!B13", f"={SN}!D75", BRL),
        ("Tributo liquido recolhido (R$)",      f"={LP}!B17", f"={LR}!B17", f"={SN}!D76", BRL),
        ("Lucro bruto do empresario (R$)",      f"={LP}!B19", f"={LR}!B19", f"={SN}!D80", BRL),
        ("Total pago pelo consumidor (R$)",     f"={LP}!B24", f"={LR}!B24", f"={SN}!D75", BRL),
        ("Carga efetiva (% do preco)",          f"={LP}!B17/{LP}!B12", f"={LR}!B17/{LR}!B13", f"={SN}!D68", PCT),
    ]
    for ri, (lbl, v_lp, v_lr, v_sn, fmt) in enumerate(antes_rows, start=20):
        rh(ws, ri, 22)
        bg = COR["VRD_CLA"] if ri % 2 == 0 else COR["CINZA"]
        cell(ws, ri, 1, lbl, bg=bg, bold=True)
        cell(ws, ri, 2, f"={v_lp}" if not v_lp.startswith("=") else v_lp, fmt=fmt, bg=bg, bold=True, halign="right")
        cell(ws, ri, 3, f"={v_lr}" if not v_lr.startswith("=") else v_lr, fmt=fmt, bg=bg, bold=True, halign="right")
        cell(ws, ri, 4, f"={v_sn}" if not v_sn.startswith("=") else v_sn, fmt=fmt, bg=bg, bold=True, halign="right")
        cell(ws, ri, 5, "", bg=bg)

    rh(ws, 25, 8)

    # ── SECAO 3: APOS A REFORMA (DEPOIS) ──────────────────────────────────
    rh(ws, 26, 22)
    secao(ws, 26, "3. APOS A REFORMA — DEPOIS (a partir de 2027)", span=5)

    rh(ws, 27, 32)
    for ci, hdr in enumerate(["Indicador", "Lucro Presumido", "Lucro Real", "Simples Nacional", ""], start=1):
        cell(ws, 27, ci, hdr, bold=True, bg=COR["MARROM"], cor_txt="FFFFFF", halign="center", sz=10)

    depois_rows = [
        ("Preco de venda / base (R$)",          f"={LP}!D12", f"={LR}!D13", f"={SN}!E75", BRL),
        ("Tributo liquido recolhido (R$)",       f"={LP}!D17", f"={LR}!D17", f"={SN}!E76", BRL),
        ("Lucro bruto do empresario (R$)",       f"={LP}!D19", f"={LR}!D19", f"={SN}!E80", BRL),
        ("Total pago pelo consumidor (R$)",      f"={LP}!D24", f"={LR}!D24", f"={SN}!E75", BRL),
        ("Carga efetiva (% do preco)",           f"={LP}!D17/{LP}!D12", f"={LR}!D17/{LR}!D13", f"={SN}!D68", PCT),
        ("CBS embutida / destacada (R$)",        f"={LP}!D12*{LP}!B7", f"={LR}!D13*{LR}!B7",
         f"={SN}!E75*{SN}!D70", BRL),
    ]
    for ri, (lbl, v_lp, v_lr, v_sn, fmt) in enumerate(depois_rows, start=28):
        rh(ws, ri, 22)
        bg = COR["AMARELO"] if ri % 2 == 0 else COR["CINZA"]
        cell(ws, ri, 1, lbl, bg=bg, bold=True)
        cell(ws, ri, 2, v_lp, fmt=fmt, bg=bg, bold=True, halign="right")
        cell(ws, ri, 3, v_lr, fmt=fmt, bg=bg, bold=True, halign="right")
        cell(ws, ri, 4, v_sn, fmt=fmt, bg=bg, bold=True, halign="right")
        cell(ws, ri, 5, "", bg=bg)

    rh(ws, 34, 8)

    # ── SECAO 4: IMPACTO DA REFORMA ───────────────────────────────────────
    rh(ws, 35, 22)
    secao(ws, 35, "4. IMPACTO DA REFORMA — VARIACAO (DEPOIS - ANTES)", span=5)

    rh(ws, 36, 32)
    for ci, hdr in enumerate(["Indicador", "Lucro Presumido", "Lucro Real", "Simples Nacional", ""], start=1):
        cell(ws, 36, ci, hdr, bold=True, bg=COR["ROXO"], cor_txt="FFFFFF", halign="center", sz=10)

    impacto_rows = [
        ("Variacao no tributo liquido (R$)",
         f"={LP}!D17-{LP}!B17", f"={LR}!D17-{LR}!B17", f"={SN}!E76-{SN}!D76", BRL,
         "Negativo = reducao de carga"),
        ("Variacao % no tributo liquido",
         f"=({LP}!D17-{LP}!B17)/{LP}!B17", f"=({LR}!D17-{LR}!B17)/{LR}!B17",
         f"=({SN}!E76-{SN}!D76)/{SN}!D76", PCT, ""),
        ("Variacao no preco ao consumidor (R$)",
         f"={LP}!D24-{LP}!B24", f"={LR}!D24-{LR}!B24", "=0", BRL,
         "SN: preco nao muda (DAS nominal inalterado)"),
        ("Variacao no lucro do empresario (R$)",
         f"={LP}!D19-{LP}!B19", f"={LR}!D19-{LR}!B19", f"={SN}!E80-{SN}!D80", BRL,
         "Positivo = ganho de margem"),
    ]
    for ri, (lbl, v_lp, v_lr, v_sn, fmt, obs) in enumerate(impacto_rows, start=37):
        rh(ws, ri, 22)
        bg = COR["LARANJA"] if ri % 2 != 0 else COR["CINZA"]
        cell(ws, ri, 1, lbl, bg=bg, bold=True)
        cell(ws, ri, 2, v_lp, fmt=fmt, bg=bg, bold=True, halign="right")
        cell(ws, ri, 3, v_lr, fmt=fmt, bg=bg, bold=True, halign="right")
        cell(ws, ri, 4, v_sn, fmt=fmt, bg=bg, bold=True, halign="right")
        cell(ws, ri, 5, obs, italic=True, sz=9, cor_txt="595959")

    rh(ws, 41, 8)

    # ── SECAO 5: RANKING RECOMENDACAO ─────────────────────────────────────
    rh(ws, 42, 22)
    secao(ws, 42, "5. COMPARATIVO E RECOMENDACAO — APOS A REFORMA", span=5)

    rh(ws, 43, 30)
    cell(ws, 43, 1, "Regime com MENOR carga tributaria (DEPOIS):", bg=COR["RESULT"], bold=True, span=2)
    f_menor = (f'=IF(AND({LP}!D17<={LR}!D17,{LP}!D17<={SN}!E76),"LUCRO PRESUMIDO",'
               f'IF(AND({LR}!D17<={LP}!D17,{LR}!D17<={SN}!E76),"LUCRO REAL","SIMPLES NACIONAL"))')
    cell(ws, 43, 3, f_menor, bg=COR["RESULT"], bold=True, sz=13, halign="center", cor_txt="375623", span=5)

    rh(ws, 44, 30)
    cell(ws, 44, 1, "Regime com MENOR preco ao consumidor (DEPOIS):", bg=COR["AZUL_CLA"], bold=True, span=2)
    f_cons = (f'=IF(AND({LP}!D24<={LR}!D24,{LP}!D24<={SN}!E75),"LUCRO PRESUMIDO",'
              f'IF(AND({LR}!D24<={LP}!D24,{LR}!D24<={SN}!E75),"LUCRO REAL","SIMPLES NACIONAL"))')
    cell(ws, 44, 3, f_cons, bg=COR["AZUL_CLA"], bold=True, sz=13, halign="center", cor_txt="1F3864", span=5)

    rh(ws, 45, 30)
    cell(ws, 45, 1, "Regime com MAIOR lucro para o empresario (DEPOIS):", bg=COR["AMARELO"], bold=True, span=2)
    f_luc = (f'=IF(AND({LP}!D19>={LR}!D19,{LP}!D19>={SN}!E80),"LUCRO PRESUMIDO",'
             f'IF(AND({LR}!D19>={LP}!D19,{LR}!D19>={SN}!E80),"LUCRO REAL","SIMPLES NACIONAL"))')
    cell(ws, 45, 3, f_luc, bg=COR["AMARELO"], bold=True, sz=13, halign="center", cor_txt="833C00", span=5)

    rh(ws, 46, 8)

    # ── SECAO 6: CRONOGRAMA DA TRANSICAO ──────────────────────────────────
    rh(ws, 47, 22)
    secao(ws, 47, "6. CRONOGRAMA DA TRANSICAO — REFORMA TRIBUTARIA (LCP 214/2025)", span=5)

    timeline = [
        ("2026", "PIS/COFINS ainda em vigor | CBS em fase de testes", COR["VRD_CLA"]),
        ("2027", "CBS comeca a substituir PIS/COFINS | aliquotas progressivas", COR["AMARELO"]),
        ("2028", "Reducao de PIS/COFINS + aumento proporcional da CBS", COR["AMARELO"]),
        ("2029", "Transicao continua — monitorar regulamentacoes do Senado", COR["AMARELO"]),
        ("2030", "Aproximacao das aliquotas definitivas da CBS", COR["LARANJA"]),
        ("2032", "Extincao total de PIS/COFINS — CBS em regime pleno", COR["LARANJA"]),
        ("2033+", "CBS consolidada | revisar precificacao e planejamento anualmente", COR["CINZA_ESC"]),
    ]
    for ri, (ano, desc, bg) in enumerate(timeline, start=48):
        rh(ws, ri, 22)
        cell(ws, ri, 1, ano, bg=bg, bold=True, halign="center", sz=11)
        cell(ws, ri, 2, desc, bg=bg, italic=True, sz=10, span=5)

    rh(ws, 55, 8)

    # ── SECAO 7: DISCLAIMER ────────────────────────────────────────────────
    rh(ws, 56, 22)
    secao(ws, 56, "7. OBSERVACOES E RESSALVAS", bg="MARROM", span=5)

    disclaimers = [
        "Este relatorio tem carater ORIENTATIVO e nao substitui a analise juridico-tributaria por profissional habilitado.",
        "Aliquotas CBS sao ESTIMATIVAS (~8,8% LP/LR | CBS embutida no DAS do SN). Serao fixadas por Resolucao do Senado Federal conforme Art.18, LCP 214/2025.",
        "Os calculos assumem formacao de preco pelo metodo MARKUP. Empresas com outras estruturas de custo devem adaptar os parametros.",
        "SIMPLES NACIONAL: A CBS entra NO LUGAR do PIS/COFINS dentro do DAS — o preco de venda NAO muda por esse motivo (aliquota nominal DAS inalterada).",
        "SPLIT PAYMENT (Arts.31-34, LCP 214): A CBS sera recolhida automaticamente pelo sistema financeiro no ato do pagamento — impacto no fluxo de caixa.",
        "Recomenda-se revisao anual do planejamento tributario ao longo do periodo de transicao (2027-2032).",
        "Fonte legal: LC 123/2006 | LC 155/2016 | LCP 214/2025 | Regulamentacao em curso pelo Senado Federal.",
    ]
    for i, d in enumerate(disclaimers):
        nota(ws, 57 + i, d)


# ═══════════════════════════════════════════════════════════════════════
#   GERAR ARQUIVO
# ═══════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()
wb.remove(wb.active)

sheet_lucro_presumido(wb)
sheet_lucro_real(wb)
sheet_simples_nacional(wb)
sheet_comparativo(wb)
sheet_regime_hibrido(wb)
sheet_relatorio_cliente(wb)

wb.active = wb["Relatorio Cliente"]

output = r"C:\Users\grazi.RAPHAEL\Downloads\Reforma_Tributaria_CBS_Comparativo_v2.xlsx"
wb.save(output)
print("Planilha criada com sucesso!")
